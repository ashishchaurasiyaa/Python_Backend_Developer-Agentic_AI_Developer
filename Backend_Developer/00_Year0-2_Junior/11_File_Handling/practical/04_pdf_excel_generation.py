"""
PDF & Excel Generation — Production Patterns
=============================================

Covers every approach taught in the theory doc:
  - WeasyPrint (HTML/CSS → PDF)        [pip install weasyprint]
  - Jinja2 template rendering           [pip install jinja2]
  - ReportLab programmatic PDF          [pip install reportlab]
  - Playwright browser-rendered PDF     [pip install playwright]
  - PyMuPDF / fitz — manipulate PDFs   [pip install pymupdf]
  - openpyxl styled workbooks           [pip install openpyxl]
  - Pandas Excel export                 [pip install pandas openpyxl xlsxwriter]
  - CSV streaming (stdlib — no install needed)
  - FastAPI streaming response          [pip install fastapi uvicorn]
  - Background generation (Celery pattern)
  - PDF parameter caching (Redis pattern)
  - Memory-efficient write_only Excel
  - i18n — currency & date formatting

All heavy external dependencies are OPTIONAL: each section gracefully
reports an ImportError when the library is absent rather than crashing.
The __main__ block at the bottom only exercises pure-stdlib sections
(CSV, i18n with stdlib locale) so the file runs without any third-party
packages installed.
"""

from __future__ import annotations

import asyncio
import csv
import hashlib
import io
import json
import locale
import logging
import uuid
import warnings
from datetime import datetime
from functools import wraps
from typing import Any, Dict, Iterator, List, Optional

log = logging.getLogger(__name__)


# ==========================================================================
# 1. DATA MODELS — shared across all generators
# ==========================================================================

class InvoiceItem:
    """A single line item on an invoice."""

    def __init__(self, name: str, qty: int, price: float) -> None:
        self.name = name
        self.qty = qty
        self.price = price

    @property
    def subtotal(self) -> float:
        return self.qty * self.price


class Invoice:
    """
    Aggregate used by every PDF/Excel generator below.

    Attributes:
        number  — human-readable invoice identifier (e.g. "INV-2024-0042")
        date    — issue date as an ISO-8601 string
        customer_name — billing entity name
        items   — list of InvoiceItem
    """

    def __init__(
        self,
        number: str,
        date: str,
        customer_name: str,
        items: Optional[List[InvoiceItem]] = None,
    ) -> None:
        self.number = number
        self.date = date
        self.customer_name = customer_name
        self.items: List[InvoiceItem] = items or []

    @property
    def total(self) -> float:
        return sum(item.subtotal for item in self.items)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "number": self.number,
            "date": self.date,
            "customer": {"name": self.customer_name},
            "items": [
                {"name": i.name, "qty": i.qty, "price": i.price}
                for i in self.items
            ],
            "total": self.total,
        }


# --------------------------------------------------------------------------
# Minimal sample dataset used throughout this module
# --------------------------------------------------------------------------

SAMPLE_INVOICE = Invoice(
    number="INV-2024-0042",
    date="2024-03-15",
    customer_name="Acme Corp",
    items=[
        InvoiceItem("Widget A", qty=3, price=49.99),
        InvoiceItem("Widget B", qty=1, price=150.00),
        InvoiceItem("Shipping", qty=1, price=12.50),
    ],
)

SALES_DATA = [
    {"date": "2024-01-01", "customer": "Alice", "amount": 100.0, "status": "Paid"},
    {"date": "2024-01-02", "customer": "Bob",   "amount": 200.0, "status": "Pending"},
    {"date": "2024-01-03", "customer": "Carol", "amount": 350.0, "status": "Paid"},
    {"date": "2024-01-04", "customer": "Dave",  "amount": 75.5,  "status": "Overdue"},
]

REFUND_DATA = [
    {"date": "2024-01-03", "customer": "Alice", "amount": 20.0, "reason": "Defect"},
    {"date": "2024-01-05", "customer": "Bob",   "amount": 50.0, "reason": "Return"},
]


# ==========================================================================
# 2. WEASYPRINT — HTML/CSS → PDF  (most common pattern)
# ==========================================================================
# pip install weasyprint

INVOICE_HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
  <style>
    body {{ font-family: Arial, sans-serif; padding: 40px; }}
    .header {{ display: flex; justify-content: space-between; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
    th, td {{ padding: 8px; border-bottom: 1px solid #ddd; text-align: left; }}
    th {{ background: #0066CC; color: #fff; }}
    .total {{ font-weight: bold; margin-top: 10px; }}
    @page {{ size: A4; margin: 1cm; }}
    @page {{
      @top-center {{ content: "My Company — Confidential"; }}
      @bottom-right {{ content: "Page " counter(page); }}
    }}
    .page-break {{ page-break-after: always; }}
  </style>
</head>
<body>
  <div class="header">
    <h1>Invoice #{number}</h1>
    <div>Date: {date}</div>
  </div>
  <h2>Bill To: {customer_name}</h2>
  <table>
    <thead><tr><th>Item</th><th>Qty</th><th>Unit Price</th><th>Subtotal</th></tr></thead>
    <tbody>{rows}</tbody>
  </table>
  <p class="total">Total: ${total:.2f}</p>
</body>
</html>
"""


def _build_invoice_html(invoice: Invoice) -> str:
    """Render the invoice HTML without Jinja2 (pure stdlib format)."""
    row_fragments = []
    for item in invoice.items:
        row_fragments.append(
            f"<tr>"
            f"<td>{item.name}</td>"
            f"<td>{item.qty}</td>"
            f"<td>${item.price:.2f}</td>"
            f"<td>${item.subtotal:.2f}</td>"
            f"</tr>"
        )
    return INVOICE_HTML_TEMPLATE.format(
        number=invoice.number,
        date=invoice.date,
        customer_name=invoice.customer_name,
        rows="".join(row_fragments),
        total=invoice.total,
    )


def generate_pdf_weasyprint(invoice: Invoice, output_path: str) -> None:
    """
    Convert an Invoice to PDF using WeasyPrint.

    WeasyPrint renders an HTML string with full CSS support (flexbox,
    @page rules, custom fonts).  This is the correct tool for 90% of
    backend PDF needs — invoices, receipts, statements.

    Args:
        invoice:     The Invoice data object.
        output_path: File path where the PDF will be written.
    """
    # pip install weasyprint
    try:
        from weasyprint import HTML  # type: ignore
    except ImportError:
        warnings.warn(
            "weasyprint not installed. "
            "Run: pip install weasyprint",
            ImportWarning,
            stacklevel=2,
        )
        return

    html_content = _build_invoice_html(invoice)
    HTML(string=html_content).write_pdf(output_path)
    log.info("WeasyPrint PDF written to %s", output_path)


# --------------------------------------------------------------------------
# 2b. Jinja2 + WeasyPrint (standard production pattern)
# --------------------------------------------------------------------------
# pip install jinja2 weasyprint

def generate_pdf_jinja2_weasyprint(
    invoice: Invoice,
    template_dir: str = "templates",
    template_name: str = "invoice.html",
) -> Optional[bytes]:
    """
    Render a Jinja2 HTML template, then convert to PDF bytes via WeasyPrint.

    This is the canonical pattern when templates are maintained separately
    by a front-end developer:
      1. Jinja2 fills in the data.
      2. WeasyPrint converts the resulting HTML to a byte stream.
      3. The caller uploads the bytes to S3 or returns them in an HTTP
         response.

    Args:
        invoice:       Data model to pass to the template.
        template_dir:  Directory that contains Jinja2 templates.
        template_name: Template filename inside template_dir.

    Returns:
        Raw PDF bytes, or None if a dependency is missing.
    """
    try:
        from jinja2 import Environment, FileSystemLoader  # type: ignore
        from weasyprint import HTML  # type: ignore
    except ImportError as exc:
        warnings.warn(
            f"Dependency missing: {exc}. "
            "Run: pip install jinja2 weasyprint",
            ImportWarning,
            stacklevel=2,
        )
        return None

    env = Environment(loader=FileSystemLoader(template_dir))
    template = env.get_template(template_name)
    html = template.render(invoice=invoice.to_dict())
    return HTML(string=html).write_pdf()


# ==========================================================================
# 3. REPORTLAB — Programmatic PDF
# ==========================================================================
# pip install reportlab

def generate_pdf_reportlab(invoice: Invoice, output_path: str) -> None:
    """
    Build a PDF programmatically using ReportLab's Platypus flow engine.

    ReportLab gives precise pixel-level control over every element.
    Use this when HTML/CSS templating does not fit (e.g. dynamic form
    fields, complex columnar layouts, embedded vector charts).

    Args:
        invoice:     The Invoice data object.
        output_path: Destination file path for the PDF.
    """
    # pip install reportlab
    try:
        from reportlab.lib import colors  # type: ignore
        from reportlab.lib.pagesizes import letter  # type: ignore
        from reportlab.lib.styles import getSampleStyleSheet  # type: ignore
        from reportlab.platypus import (  # type: ignore
            Paragraph,
            SimpleDocTemplate,
            Table,
            TableStyle,
        )
    except ImportError:
        warnings.warn(
            "reportlab not installed. "
            "Run: pip install reportlab",
            ImportWarning,
            stacklevel=2,
        )
        return

    doc = SimpleDocTemplate(output_path, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # --- Title and metadata ---
    story.append(Paragraph(f"Invoice #{invoice.number}", styles["Title"]))
    story.append(Paragraph(f"Date: {invoice.date}", styles["Normal"]))
    story.append(Paragraph(f"Bill To: {invoice.customer_name}", styles["Normal"]))

    # --- Line items table ---
    table_data: List[List[str]] = [["Item", "Qty", "Unit Price", "Subtotal"]]
    for item in invoice.items:
        table_data.append(
            [item.name, str(item.qty), f"${item.price:.2f}", f"${item.subtotal:.2f}"]
        )
    table_data.append(["", "", "Total", f"${invoice.total:.2f}"])

    table = Table(table_data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                # Header row styling
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0066CC")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                # Grid lines
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                # Totals row — bold
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F0F0F0")),
                # Cell alignment
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    log.info("ReportLab PDF written to %s", output_path)


# ==========================================================================
# 4. PLAYWRIGHT — Browser-rendered PDF (highest CSS fidelity)
# ==========================================================================
# pip install playwright && playwright install chromium

async def generate_pdf_playwright(html_content: str, output_path: str) -> None:
    """
    Launch a headless Chromium browser, load HTML, and export to PDF.

    Use this when WeasyPrint cannot handle the required CSS (animations,
    JS-rendered charts, shadow DOM).  Slower than WeasyPrint (~2-5 s per
    page) but produces pixel-perfect output.

    Args:
        html_content: Full HTML document as a string.
        output_path:  Destination file path for the PDF.
    """
    # pip install playwright && playwright install chromium
    try:
        from playwright.async_api import async_playwright  # type: ignore
    except ImportError:
        warnings.warn(
            "playwright not installed. "
            "Run: pip install playwright && playwright install chromium",
            ImportWarning,
            stacklevel=2,
        )
        return

    async with async_playwright() as p:
        browser = await p.chromium.launch()
        page = await browser.new_page()
        await page.set_content(html_content, wait_until="networkidle")
        await page.pdf(
            path=output_path,
            format="A4",
            print_background=True,
            margin={"top": "1cm", "bottom": "1cm", "left": "1cm", "right": "1cm"},
        )
        await browser.close()

    log.info("Playwright PDF written to %s", output_path)


# ==========================================================================
# 5. PyMuPDF / fitz — Manipulate Existing PDFs
# ==========================================================================
# pip install pymupdf

def extract_text_from_pdf(input_path: str) -> List[str]:
    """
    Open an existing PDF and return the text content of each page.

    Args:
        input_path: Path to the source PDF file.

    Returns:
        A list of strings, one per page.  Returns empty list if
        pymupdf is not installed.
    """
    # pip install pymupdf
    try:
        import fitz  # type: ignore  # PyMuPDF
    except ImportError:
        warnings.warn(
            "pymupdf not installed. "
            "Run: pip install pymupdf",
            ImportWarning,
            stacklevel=2,
        )
        return []

    doc = fitz.open(input_path)
    pages_text = [page.get_text() for page in doc]
    log.info("Extracted text from %d pages in %s", doc.page_count, input_path)
    return pages_text


def merge_pdfs(input_paths: List[str], output_path: str) -> None:
    """
    Merge multiple PDF files into a single document.

    Args:
        input_paths: Ordered list of PDF file paths to merge.
        output_path: Destination path for the merged PDF.
    """
    try:
        import fitz  # type: ignore
    except ImportError:
        warnings.warn("pymupdf not installed.", ImportWarning, stacklevel=2)
        return

    merged = fitz.open()
    for path in input_paths:
        merged.insert_pdf(fitz.open(path))
    merged.save(output_path)
    log.info("Merged %d PDFs → %s", len(input_paths), output_path)


def split_pdf(input_path: str, output_dir: str) -> None:
    """
    Split a multi-page PDF into one file per page.

    Args:
        input_path: Source PDF to split.
        output_dir: Directory where individual page PDFs will be saved.
    """
    try:
        import fitz  # type: ignore
    except ImportError:
        warnings.warn("pymupdf not installed.", ImportWarning, stacklevel=2)
        return

    doc = fitz.open(input_path)
    for i, _ in enumerate(doc):
        single = fitz.open()
        single.insert_pdf(doc, from_page=i, to_page=i)
        single.save(f"{output_dir}/page_{i + 1}.pdf")
    log.info("Split %s into %d pages in %s", input_path, doc.page_count, output_dir)


def add_watermark(input_path: str, output_path: str, text: str = "CONFIDENTIAL") -> None:
    """
    Stamp a text watermark on every page of a PDF.

    Args:
        input_path:  Source PDF.
        output_path: Watermarked PDF destination.
        text:        Watermark string to insert.
    """
    try:
        import fitz  # type: ignore
    except ImportError:
        warnings.warn("pymupdf not installed.", ImportWarning, stacklevel=2)
        return

    doc = fitz.open(input_path)
    for page in doc:
        rect = fitz.Rect(50, 50, 300, 120)
        page.insert_textbox(
            rect,
            text,
            fontsize=36,
            color=(0.8, 0.0, 0.0),  # Red in RGB
            rotate=30,
        )
    doc.save(output_path)
    log.info("Watermark '%s' applied → %s", text, output_path)


# ==========================================================================
# 6. openpyxl — Styled Excel Workbook
# ==========================================================================
# pip install openpyxl

def generate_excel_openpyxl(
    sales: List[Dict[str, Any]],
    output_path: str,
) -> None:
    """
    Create a styled .xlsx workbook from a list of sales records.

    Features demonstrated:
      - Header row with bold white text on a blue background
      - Centred alignment for all header cells
      - Auto-sized column widths
      - SUM formula injected into the Amount column footer
      - A bar chart anchored to cell E2

    Args:
        sales:       List of dicts with keys: date, customer, amount, status.
        output_path: Destination .xlsx file path.
    """
    # pip install openpyxl
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.chart import BarChart, Reference  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError:
        warnings.warn(
            "openpyxl not installed. "
            "Run: pip install openpyxl",
            ImportWarning,
            stacklevel=2,
        )
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"

    # --- Header row ---
    headers = ["Date", "Customer", "Amount", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0066CC")
        cell.alignment = Alignment(horizontal="center")

    # --- Data rows ---
    for record in sales:
        ws.append([
            record["date"],
            record["customer"],
            record["amount"],
            record["status"],
        ])

    # --- Auto-size columns ---
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # --- SUM formula in footer ---
    data_end_row = len(sales) + 1  # header is row 1, data starts at row 2
    ws.cell(
        row=data_end_row + 1,
        column=3,
        value=f"=SUM(C2:C{data_end_row})",
    ).font = Font(bold=True)

    # --- Bar chart on Amount column ---
    chart = BarChart()
    chart.title = "Sales Amount"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = "Row"
    data_ref = Reference(
        ws, min_col=3, min_row=1, max_col=3, max_row=data_end_row
    )
    chart.add_data(data_ref, titles_from_data=True)
    ws.add_chart(chart, "F2")

    wb.save(output_path)
    log.info("openpyxl workbook written to %s", output_path)


# ==========================================================================
# 7. openpyxl write_only mode — Memory-efficient large exports
# ==========================================================================
# pip install openpyxl

def generate_excel_large_dataset(
    rows_iter: Iterator[List[Any]],
    headers: List[str],
    output_path: str,
) -> None:
    """
    Stream millions of rows into an Excel file without loading them all
    into memory by using openpyxl's write_only mode.

    In write_only mode the workbook never holds a full in-memory grid;
    rows are serialised directly to the zip stream.  This is the correct
    approach for exports > 100 K rows.

    Args:
        rows_iter:   An iterator that yields one row (list of values) at a time.
        headers:     Column header labels.
        output_path: Destination .xlsx file path.
    """
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError:
        warnings.warn("openpyxl not installed.", ImportWarning, stacklevel=2)
        return

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Export")
    ws.append(headers)
    for row in rows_iter:
        ws.append(row)
    wb.save(output_path)
    log.info("write_only workbook written to %s", output_path)


# ==========================================================================
# 8. Pandas — Tabular data to Excel / multiple sheets
# ==========================================================================
# pip install pandas openpyxl

def generate_excel_pandas(
    sales: List[Dict[str, Any]],
    refunds: List[Dict[str, Any]],
    output_path: str,
) -> None:
    """
    Export two DataFrames (Sales and Refunds) to separate sheets in one
    .xlsx file using pandas ExcelWriter.

    pandas to_excel is the fastest route when the data is already in a
    DataFrame — it requires no manual cell iteration.

    Args:
        sales:       List of sales records (will become sheet "Sales").
        refunds:     List of refund records (will become sheet "Refunds").
        output_path: Destination .xlsx file path.
    """
    # pip install pandas openpyxl
    try:
        import pandas as pd  # type: ignore
    except ImportError:
        warnings.warn(
            "pandas not installed. "
            "Run: pip install pandas openpyxl",
            ImportWarning,
            stacklevel=2,
        )
        return

    df_sales = pd.DataFrame(sales)
    df_refunds = pd.DataFrame(refunds)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_sales.to_excel(writer, sheet_name="Sales", index=False)
        df_refunds.to_excel(writer, sheet_name="Refunds", index=False)

    log.info("Pandas multi-sheet workbook written to %s", output_path)


# ==========================================================================
# 9. CSV — Streaming export for large datasets (stdlib only)
# ==========================================================================

def write_csv(
    records: List[Dict[str, Any]],
    fieldnames: List[str],
    output_path: str,
) -> None:
    """
    Write a list of dicts to a CSV file using the stdlib csv module.

    For exports > 100 K rows, CSV is faster and smaller than Excel.
    The stdlib csv module needs no third-party dependencies.

    Args:
        records:    Iterable of flat dicts.
        fieldnames: Column order; must match dict keys.
        output_path: Destination .csv file path.
    """
    with open(output_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)
    log.info("CSV written to %s (%d rows)", output_path, len(records))


# ==========================================================================
# 10. FastAPI — Streaming CSV response
# ==========================================================================
# pip install fastapi uvicorn

# NOTE: This function is illustrative.  It references FastAPI's
# StreamingResponse and an async row iterator pattern that only makes
# sense inside a running ASGI app.

async def _stream_csv_bytes(rows_iter: Any) -> Any:
    """
    Async generator that yields CSV bytes row by row.

    Memory footprint is constant regardless of result set size because
    only one row is buffered at a time.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["id", "customer", "amount"])
    yield buffer.getvalue().encode()
    buffer.seek(0)
    buffer.truncate(0)

    async for row in rows_iter:
        writer.writerow(row)
        yield buffer.getvalue().encode()
        buffer.seek(0)
        buffer.truncate(0)


def make_streaming_csv_response(rows_iter: Any) -> Any:
    """
    Return a FastAPI StreamingResponse for a large CSV export.

    This keeps API memory usage constant even for millions of rows.
    The Content-Disposition header triggers a file-download in browsers.

    Args:
        rows_iter: Async iterator of row tuples from a DB cursor.

    Returns:
        A FastAPI StreamingResponse object, or None if fastapi is absent.
    """
    # pip install fastapi
    try:
        from fastapi.responses import StreamingResponse  # type: ignore
    except ImportError:
        warnings.warn(
            "fastapi not installed. "
            "Run: pip install fastapi",
            ImportWarning,
            stacklevel=2,
        )
        return None

    return StreamingResponse(
        _stream_csv_bytes(rows_iter),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=orders.csv"},
    )


# ==========================================================================
# 11. Background generation pattern (Celery + S3)
# ==========================================================================

# NOTE: This section illustrates the architecture described in the theory
# doc.  The Celery/S3 calls are commented out because they require running
# infrastructure (broker, bucket).  The orchestration logic is shown
# without those calls so the file remains syntax-clean.

def request_report_generation(report_params: Dict[str, Any]) -> str:
    """
    Enqueue a background report-generation job and return a job_id.

    In production this function:
      1. Writes a 'queued' row to the database.
      2. Dispatches a Celery task asynchronously.
      3. Returns the job_id immediately so the API responds in < 1 ms.

    The slow PDF/Excel work happens in a separate worker process.

    Args:
        report_params: Arbitrary dict describing what to generate.

    Returns:
        A UUID job_id string the caller can poll for status.
    """
    job_id = str(uuid.uuid4())

    # --- PRODUCTION ---
    # await db.execute(
    #     "INSERT INTO reports (id, status) VALUES ($1, 'queued')",
    #     job_id,
    # )
    # generate_report_celery_task.delay(job_id, report_params)

    log.info("Report job %s enqueued with params: %s", job_id, report_params)
    return job_id


def generate_report_worker(job_id: str, params: Dict[str, Any]) -> None:
    """
    Celery worker body: generate the file, upload to S3, notify user.

    This runs in a separate process (worker), so heavy memory usage from
    PDF generation does not affect the API process.

    Args:
        job_id: Corresponds to a row in the reports table.
        params: Same dict that was passed to request_report_generation.
    """
    # In production, the @celery.task decorator would wrap this function.

    # Step 1 — Generate bytes
    # pdf_bytes = generate_pdf_weasyprint(...)

    # Step 2 — Upload to S3
    # s3.put_object(Bucket=BUCKET, Key=f"reports/{job_id}.pdf", Body=pdf_bytes)

    # Step 3 — Mark done in database
    # db.execute("UPDATE reports SET status='done' WHERE id=$1", job_id)

    # Step 4 — Notify user
    # send_email(user_email, "Your report is ready", ...)

    log.info("Worker stub for job %s — no infrastructure connected.", job_id)


# ==========================================================================
# 12. PDF caching — hash parameters, cache bytes in Redis
# ==========================================================================

async def get_or_generate_pdf_cached(
    params: Dict[str, Any],
    redis_client: Any,
    ttl_seconds: int = 3600,
) -> bytes:
    """
    Return cached PDF bytes if available; generate and cache otherwise.

    The cache key is the SHA-256 digest of the JSON-serialised params
    (sorted keys guarantee identical dicts map to the same key).
    This prevents regenerating the same PDF when many users request
    identical report parameters.

    Args:
        params:        Report parameters (must be JSON-serialisable).
        redis_client:  An async Redis client (e.g. redis.asyncio.Redis).
        ttl_seconds:   Cache TTL; defaults to 1 hour.

    Returns:
        Raw PDF bytes.
    """
    cache_key = "pdf:" + hashlib.sha256(
        json.dumps(params, sort_keys=True).encode()
    ).hexdigest()

    # Try cache first
    cached: Optional[bytes] = await redis_client.get(cache_key)
    if cached:
        log.info("Cache hit for key %s", cache_key)
        return cached

    # Generate (expensive)
    # pdf_bytes = await generate_pdf(params)
    pdf_bytes = b"%PDF placeholder"  # Replace with real generator call

    await redis_client.setex(cache_key, ttl_seconds, pdf_bytes)
    log.info("Cached PDF under key %s (TTL=%ds)", cache_key, ttl_seconds)
    return pdf_bytes


# ==========================================================================
# 13. i18n — Currency and date formatting
# ==========================================================================

def format_currency_locale(amount: float, locale_name: str = "en_US.UTF-8") -> str:
    """
    Format a numeric amount as a locale-aware currency string.

    Example:
        format_currency_locale(100000, "en_IN.UTF-8")
        # Returns "₹1,00,000.00" (Indian format)

    Args:
        amount:      Numeric value to format.
        locale_name: POSIX locale string (e.g. "en_US.UTF-8", "en_IN.UTF-8").

    Returns:
        Formatted currency string, or a plain fallback if the locale is
        unavailable on this system.
    """
    try:
        locale.setlocale(locale.LC_ALL, locale_name)
        return locale.currency(amount, grouping=True)
    except locale.Error:
        # Locale not installed on this OS; fall back to plain formatting
        return f"${amount:,.2f}"
    finally:
        # Restore default locale to avoid side-effects on other formatters
        locale.setlocale(locale.LC_ALL, "")


def format_date_locale(dt: datetime, locale_name: str, fmt: str = "%A, %d %B %Y") -> str:
    """
    Format a datetime object using a locale-specific date representation.

    Example:
        format_date_locale(datetime.now(), "es_ES.UTF-8")
        # Returns "viernes, 15 marzo 2024" on macOS (Spanish)

    Args:
        dt:          The datetime to format.
        locale_name: POSIX locale string.
        fmt:         strftime format string.

    Returns:
        Locale-formatted date string, or ISO-8601 fallback.
    """
    try:
        locale.setlocale(locale.LC_TIME, locale_name)
        return dt.strftime(fmt)
    except locale.Error:
        return dt.strftime("%Y-%m-%d")
    finally:
        locale.setlocale(locale.LC_TIME, "")


# ==========================================================================
# 14. File naming convention helper
# ==========================================================================

def make_report_filename(prefix: str, extension: str = "pdf") -> str:
    """
    Generate a collision-resistant report filename.

    Format: {prefix}_{YYYYMMDD_HHMMSS}_{short_uuid}.{extension}

    Example:
        make_report_filename("invoice")
        # Returns "invoice_20240315_143022_a1b2c3d4.pdf"

    Avoids the common pitfall of same-named files overwriting each other
    when many reports are generated concurrently.

    Args:
        prefix:    Human-readable identifier (e.g. "invoice", "sales_report").
        extension: File extension without dot (default "pdf").

    Returns:
        A unique filename string.
    """
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    short_id = str(uuid.uuid4()).replace("-", "")[:8]
    return f"{prefix}_{timestamp}_{short_id}.{extension}"


# ==========================================================================
# 15. Audit log helper
# ==========================================================================

def audit_log_report_generation(
    user_id: str,
    report_type: str,
    filename: str,
    params: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Produce an audit log entry for a report generation event.

    Compliance-regulated backends (finance, healthcare) must record who
    requested a file, when, and with what parameters so that a full
    paper trail exists.

    Args:
        user_id:     Authenticated user identifier.
        report_type: E.g. "invoice_pdf", "sales_xlsx".
        filename:    The generated filename (see make_report_filename).
        params:      Sanitised parameters (strip PII if required).

    Returns:
        A dict suitable for insertion into an audit_log table or
        structured logging pipeline.
    """
    entry: Dict[str, Any] = {
        "event": "report_generated",
        "user_id": user_id,
        "report_type": report_type,
        "filename": filename,
        "params": params,
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "correlation_id": str(uuid.uuid4()),
    }
    log.info("AUDIT %s", json.dumps(entry))
    return entry


# ==========================================================================
# 16. ProcessPool isolation for heavy PDF work
# ==========================================================================

def _generate_pdf_subprocess(params: Dict[str, Any]) -> bytes:
    """
    Worker function to be executed in a subprocess via ProcessPoolExecutor.

    WeasyPrint can consume 500+ MB for complex documents.  Running it in
    a subprocess isolates that memory spike from the main API process;
    the OS reclaims all memory when the subprocess exits.

    Args:
        params: Dict with 'invoice' key (serialised invoice data).

    Returns:
        Raw PDF bytes (or a placeholder in this illustration).
    """
    # In production: rebuild the Invoice from params and call WeasyPrint.
    # invoice = Invoice(**params["invoice"])
    # html = _build_invoice_html(invoice)
    # from weasyprint import HTML
    # return HTML(string=html).write_pdf()
    log.info("Subprocess PDF generation for params: %s", params)
    return b"%PDF-1.4 subprocess placeholder"


async def generate_pdf_in_subprocess(params: Dict[str, Any]) -> bytes:
    """
    Run PDF generation in a separate process to protect the event loop.

    asyncio.get_event_loop().run_in_executor with a ProcessPoolExecutor
    offloads CPU/memory-intensive work without blocking the ASGI event
    loop.

    Args:
        params: Serialisable parameters forwarded to the subprocess worker.

    Returns:
        Raw PDF bytes from the subprocess.
    """
    import concurrent.futures

    loop = asyncio.get_event_loop()
    with concurrent.futures.ProcessPoolExecutor(max_workers=2) as pool:
        result: bytes = await loop.run_in_executor(
            pool, _generate_pdf_subprocess, params
        )
    return result


# ==========================================================================
# 17. PITFALLS REFERENCE (commented catalogue)
# ==========================================================================

PITFALLS = """
Common pitfalls when generating PDFs and Excel files in production:

1. GENERATING PDF IN AN API REQUEST
   - PDF generation can take 1–30 seconds.
   - Blocking the request thread starves other clients.
   - Fix: dispatch to a background queue (Celery, ARQ, RQ).

2. LOADING 1M-ROW EXCEL IN MEMORY
   - openpyxl reads the entire workbook into RAM.
   - Fix: use write_only=True mode or export as CSV instead.

3. MISSING / INCORRECTLY BUNDLED FONTS
   - Custom @font-face fonts not shipped with the app.
   - WeasyPrint silently falls back to an ugly default typeface.
   - Fix: embed font files in the Docker image; reference them by path.

4. RGB vs CMYK COLOR SPACE
   - Print PDFs require CMYK; screen PDFs use RGB.
   - WeasyPrint defaults to RGB — fine for screen/email, wrong for press.
   - Fix: for press-ready PDFs use a dedicated PDF library with CMYK support.

5. NO FILE NAMING CONVENTION
   - Files with the same name silently overwrite each other in S3/disk.
   - Fix: include timestamp + UUID in every filename (see make_report_filename).

6. EXPOSING INTERNAL DATA IN PDFs
   - Rendering a template that includes internal_id, cost_price, margin.
   - Fix: strip sensitive fields in the data layer before passing to template.

7. SYNCHRONOUS IMAGE FETCH INSIDE TEMPLATES
   - Each <img> tag triggers a separate HTTP request during rendering.
   - Serialises badly at scale; one slow CDN call stalls the whole PDF.
   - Fix: pre-download images; embed them as base64 data URIs in the template.
"""


# ==========================================================================
# __main__ — runs without any third-party libraries
# ==========================================================================

if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
    )

    print("\n" + "=" * 70)
    print("  PDF & Excel Generation — Practical Demo (stdlib-only sections)")
    print("=" * 70)

    # --- Section 9: CSV export ---
    print("\n[Section 9] CSV export")
    csv_path = "/tmp/sales_demo.csv"
    write_csv(
        records=SALES_DATA,
        fieldnames=["date", "customer", "amount", "status"],
        output_path=csv_path,
    )
    print(f"  Written: {csv_path}")

    # --- Section 13: i18n currency ---
    print("\n[Section 13] Currency formatting")
    for loc, amount in [("en_US.UTF-8", 1500.75), ("en_IN.UTF-8", 100000)]:
        result = format_currency_locale(amount, loc)
        print(f"  locale={loc!r:<20} amount={amount:<12} → {result}")

    # --- Section 13: i18n date ---
    print("\n[Section 13] Date formatting")
    now = datetime(2024, 3, 15)
    for loc in ["en_US.UTF-8", "es_ES.UTF-8", "fr_FR.UTF-8"]:
        result = format_date_locale(now, loc)
        print(f"  locale={loc!r:<20} → {result}")

    # --- Section 14: filename convention ---
    print("\n[Section 14] Report filename generation")
    for prefix, ext in [("invoice", "pdf"), ("sales_report", "xlsx"), ("orders", "csv")]:
        print(f"  {make_report_filename(prefix, ext)}")

    # --- Section 15: audit log ---
    print("\n[Section 15] Audit log entry")
    entry = audit_log_report_generation(
        user_id="user_001",
        report_type="invoice_pdf",
        filename=make_report_filename("invoice"),
        params={"invoice_number": "INV-2024-0042"},
    )
    print(f"  correlation_id={entry['correlation_id']}")

    # --- Section 11: background job stub ---
    print("\n[Section 11] Background job enqueue stub")
    job_id = request_report_generation({"type": "sales", "month": "2024-01"})
    print(f"  Enqueued job_id={job_id}")

    # --- Show invoice model total ---
    print("\n[Data model] Sample invoice total")
    inv = SAMPLE_INVOICE
    print(f"  Invoice #{inv.number}  Customer: {inv.customer_name}")
    for item in inv.items:
        print(f"    {item.name:<15} qty={item.qty}  ${item.subtotal:.2f}")
    print(f"  TOTAL: ${inv.total:.2f}")

    print("\n[INFO] To test PDF/Excel generators install the optional deps:")
    print("  pip install weasyprint reportlab openpyxl pandas pymupdf jinja2")
    print("=" * 70 + "\n")
